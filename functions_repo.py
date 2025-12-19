# -------------------------
# -------------------------
# Utilities
# -------------------------
# -------------------------
# ---- Standard library
import os
import sys
import math
import re
import random
import warnings
from math import sqrt
from typing import Any, Optional, Sequence, Dict, Union, Callable, Tuple
from collections import defaultdict, Counter
from pathlib import Path

# ---- Core numerics & utils
import numpy as np
import pandas as pd
import shap
from joblib import Parallel, delayed
from scipy.stats import (
    ttest_ind,
    ttest_ind_from_stats,
    mannwhitneyu,
    pearsonr,
    spearmanr,
    truncnorm,
    f_oneway,
)
from statsmodels.stats.multitest import multipletests
from statsmodels.stats.multicomp import pairwise_tukeyhsd
import statsmodels.api as sm

# ---- Plotting
import matplotlib.pyplot as plt
from matplotlib.ticker import MaxNLocator
import seaborn as sns
import matplotlib.patches as mpatches
from matplotlib.patches import Rectangle
import matplotlib as mpl
from matplotlib.lines import Line2D

# ---- Excel I/O
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage  # Pillow-backed

# ---- Scikit-learn
from sklearn.base import clone, BaseEstimator, TransformerMixin
from sklearn.dummy import DummyRegressor
from sklearn.decomposition import PCA
from sklearn.cross_decomposition import PLSRegression
from sklearn.feature_selection import SelectKBest, f_regression
from sklearn.linear_model import ElasticNetCV, Ridge
from sklearn.metrics import mean_squared_error, mean_absolute_error, r2_score
from sklearn.model_selection import LeaveOneOut, GroupShuffleSplit
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler, FunctionTransformer
from sklearn.cluster import AffinityPropagation
from sklearn.metrics.pairwise import cosine_similarity, pairwise_distances

# ---- PyTorch / skorch (neural nets)
import torch
import torch.nn as nn
from skorch import NeuralNetRegressor
from skorch.dataset import ValidSplit, Dataset
from skorch.callbacks import LRScheduler, GradientNormClipping, EarlyStopping
from skorch.helper import predefined_split

# ---- Optuna (HPO)
import optuna
from optuna.samplers import TPESampler

# ---- Project helpers
#from functions_repo import *

# ============================================
# Reproducibility & GPU-determinism hints
# ============================================
RANDOM_STATE = 42
# Ensure deterministic CUDA math (where applicable)
os.environ.setdefault("PYTHONHASHSEED", "42")
os.environ.setdefault("CUBLAS_WORKSPACE_CONFIG", ":4096:8")
os.environ.setdefault("NVIDIA_TF32_OVERRIDE", "0")
# Local seed utility (kept to preserve original behavior; later import of set_seed
# from functions_repo intentionally shadows this definition, as in the source code.)
def set_seed(seed=42):
    # RNG seeds
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)
    # Deterministic kernels (raises if a non-deterministic op is used)
    torch.use_deterministic_algorithms(True)
    # cuDNN: no autotune; fixed choices
    torch.backends.cudnn.benchmark = False
    torch.backends.cudnn.deterministic = True
    # Turn TF32 OFF everywhere (important on A100/H100; harmless on V100)
    torch.backends.cuda.matmul.allow_tf32 = False
    torch.backends.cudnn.allow_tf32 = False
    # Prefer stable FP32 matmul route (avoid "highest" which may pick TF32-like paths)
    try:
        torch.set_float32_matmul_precision("highest")
    except Exception:
        pass
    # (Optional) stabilize CPU math if you ever eval on CPU
    os.environ["PYTHONHASHSEED"] = str(seed)
    os.environ["CUBLAS_WORKSPACE_CONFIG"] = ":4096:8"
    os.environ.setdefault("OMP_NUM_THREADS", "32")
    os.environ.setdefault("MKL_NUM_THREADS", "32")
    os.environ.setdefault("OPENBLAS_NUM_THREADS", "32")
    try:
        torch.set_num_threads(int(os.environ.get("OMP_NUM_THREADS", "32")))
    except Exception:
        pass
    # Guardrails: hard-assert nothing flipped TF32 back on
    assert not torch.backends.cuda.matmul.allow_tf32, "TF32 still ON for cuBLAS!"
    assert not torch.backends.cudnn.allow_tf32, "TF32 still ON for cuDNN!"

set_seed(42)


def _canon_acc2kegg(acc2kegg_df: pd.DataFrame) -> pd.DataFrame:
    """Standardize AccessionToKEGG to columns: UniProt, KEGG."""
    df = acc2kegg_df.copy()
    rename = {}
    for c in df.columns:
        cl = str(c).lower()
        if "uniprot" in cl or "accession" in cl:
            rename[c] = "UniProt"
        elif "kegg" in cl:
            rename[c] = "KEGG"
    if rename:
        df = df.rename(columns=rename)
    if "UniProt" not in df.columns or "KEGG" not in df.columns:
        if len(df.columns) >= 2:
            df = df.rename(columns={df.columns[0]: "UniProt", df.columns[1]: "KEGG"})
        else:
            raise ValueError("AccessionToKEGG must have at least two columns (UniProt, KEGG).")
    df["UniProt"] = df["UniProt"].astype(str).str.strip()
    df["KEGG"]    = df["KEGG"].astype(str).str.strip()
    df = df.dropna(subset=["UniProt"]).drop_duplicates()
    return df[["UniProt","KEGG"]].reset_index(drop=True)

def _build_to_uniprot_map(acc2kegg_df: pd.DataFrame) -> dict:
    """Dict that maps either UniProt or KEGG -> UniProt."""
    df = _canon_acc2kegg(acc2kegg_df)
    m = {}
    for _, r in df.iterrows():
        up = str(r["UniProt"]).strip()
        kg = str(r["KEGG"]).strip()
        if up:
            m[up] = up
        if kg and kg.lower() != "nan":
            m[kg] = up
    return m


def loocv_with_anchor_indices(cond_keys, always_train_keys):
    """
    Yields (train_idx, test_idx) where test is the FULL set of indices
    for one non-anchor Condition_key (i.e., all 5 replicates).
    """
    N = len(cond_keys)
    # All unique condition groups present in the data
    all_groups = pd.unique(cond_keys)
    # Candidate test groups = all groups except anchors (anchors must always remain in train)
    candidate_test_groups = [g for g in all_groups if g not in always_train_keys]
    for g in candidate_test_groups:
        # indices for the held-out test condition (all replicates)
        test_idx = np.where(cond_keys == g)[0]
        # everything else is train
        train_idx = np.setdiff1d(np.arange(N), test_idx)
        # sanity: anchors are not in test and remain in train
        assert not np.any(np.isin(cond_keys[test_idx], list(always_train_keys))), "Anchor leaked into test!"
        anchor_idx = np.where(np.isin(cond_keys, list(always_train_keys)))[0]
        assert np.all(np.isin(anchor_idx, train_idx)), "Anchor missing from train!"
        yield train_idx, test_idx

# Helpful for the percentage-accuracy metric to avoid division-by-zero
EPS = 1e-12

def augment_labels_normal(X_train, y_train, K=5, rng=42, min_gr=0.0, truncated=True, eps=1e-12, return_indices=False):
    """
    X_train : (N, d)
    y_train : (N, 2) with [:,0]=GR_mean (mu), [:,1]=GR_sd (sd)

    Returns:
      X_all : (N + N*K, d)  -> originals first, then all augmented rows
      y_all : (N + N*K, 2)  -> for augmented: [:,0] are sampled GRs, [:,1] keeps sd
      (opt) idx_orig : slice or array of indices for originals (0..N-1)
      (opt) idx_aug  : slice or array of indices for augmented (N..N+N*K-1)
    """
    rs = np.random.RandomState(rng)

    X_train = np.asarray(X_train, dtype=np.float32)
    y_train = np.asarray(y_train, dtype=np.float64)
    mu = y_train[:, 0]
    sd = np.maximum(y_train[:, 1], eps)  # avoid sd=0
    N = mu.shape[0]

    if K <= 0:
        # nothing to augment; just return originals
        if return_indices:
            return X_train, y_train.astype(np.float32), np.arange(N), np.array([], dtype=int)
        return X_train, y_train.astype(np.float32)

    if truncated:
        # vectorized truncated normal: shape (N, K)
        a = ((min_gr - mu) / sd)[:, None]                 # (N,1)
        b = np.full((N, 1), np.inf, dtype=np.float64)     # (N,1)
        loc = mu[:, None]                                 # (N,1)
        scale = sd[:, None]                               # (N,1)
        Ys = truncnorm.rvs(a, b, loc=loc, scale=scale, size=(N, K), random_state=rs)
    else:
        # plain normal, then clamp at min_gr
        noise = rs.normal(loc=0.0, scale=sd[:, None], size=(N, K))  # (N,K)
        Ys = np.maximum(mu[:, None] + noise, min_gr)

    # Build augmented block (all augmented AFTER originals)
    X_aug = np.repeat(X_train, K, axis=0)  # (N*K, d)
    y_aug = np.stack([Ys.ravel(), np.repeat(sd, K)], axis=1).astype(np.float32)  # (N*K, 2)

    # Concatenate: originals first, augmented after
    X_all = np.vstack([X_train, X_aug]).astype(np.float32)
    y_all = np.vstack([y_train.astype(np.float32), y_aug]).astype(np.float32)

    if return_indices:
        idx_orig = np.arange(0, N, dtype=int)
        idx_aug  = np.arange(N, N + N*K, dtype=int)
        return X_all, y_all, idx_orig, idx_aug
    return X_all, y_all

# === extra utilities ===
def _internal_split_sizes(n_train_fold, valid_frac=0.2):
    # Mirrors ValidSplit(0.2) rounding and guarantees at least 1 valid sample
    n_valid = max(1, int(round(valid_frac * n_train_fold)))
    n_train_internal = n_train_fold - n_valid
    return n_train_internal, n_valid

def percent_accuracy(y_true, y_pred, eps=1e-12):
    y_true = np.asarray(y_true).ravel()
    y_pred = np.asarray(y_pred).ravel()
    den = np.maximum(np.abs(y_true), eps)           # elementwise
    rel_err = np.abs((y_true - y_pred) / den)
    return (1.0 - np.mean(rel_err)) * 100.0

def percent_accuracy_scalar(y_true_scalar, y_pred_scalar, eps=1e-12):
    den = max(abs(float(y_true_scalar)), eps)       # scalar
    rel_err = abs(float(y_true_scalar) - float(y_pred_scalar)) / den
    return (1.0 - rel_err) * 100.0

def smape(y_true, y_pred, eps=1e-12):
    y_true = np.asarray(y_true, dtype=float).ravel()
    y_pred = np.asarray(y_pred, dtype=float).ravel()
    denom = (np.abs(y_true) + np.abs(y_pred)) / 2.0
    denom = np.maximum(denom, eps)  # avoid 0/0
    return 100.0 * np.mean(np.abs(y_true - y_pred) / denom)

# --------- pick k random non-anchor conditions for validation (single split) ----------
def make_random_group_valid(train_idx, cond_keys, always_train_keys, k=2, rng=42):
    """
    From the outer train set, choose k condition groups (non-anchors) as validation.
    Returns (inner_train_idx, inner_valid_idx, chosen_groups).
    Assumes: global 'cond_keys' and 'always_train_keys' are defined.
    """
    rs = np.random.RandomState(rng)

    train_groups = cond_keys[train_idx]
    non_anchor_groups = np.unique(train_groups[~np.isin(train_groups, list(always_train_keys))])

    if len(non_anchor_groups) == 0:
        # No eligible groups -> no validation
        return train_idx, np.array([], dtype=int), []

    k_eff = min(k, len(non_anchor_groups))
    chosen = rs.choice(non_anchor_groups, size=k_eff, replace=False)

    valid_mask = np.isin(cond_keys[train_idx], chosen)
    inner_valid_idx = train_idx[valid_mask]
    inner_train_idx = np.setdiff1d(train_idx, inner_valid_idx)

    return inner_train_idx, inner_valid_idx, chosen.tolist()

# --------- LOOCV with single random-k-group validation per outer fold ----------
def loocv_scores(pipe, X, y,  cond_keys, always_train_keys, k_valid_groups=2, rng=42, K_AUG=200):
    mse_scores, rmse_scores, mae_scores, acc_scores, smape_scores = [], [], [], [], []
    y_pred_all = np.zeros(y.shape[0], dtype=float)
    fold_info = []

    # ----- condition-level accumulators -----
    cond_mse_scores, cond_rmse_scores, cond_mae_scores = [], [], []
    cond_pred_dict, cond_true_dict = {}, {}

    def _cv(a, eps=1e-12):
        a = np.asarray(a, float)
        if a.size <= 1:
            return 0.0
        m = float(np.mean(a))
        s = float(np.std(a, ddof=1))
        return s / (m + eps)

    for fold_idx, (train_idx, test_idx) in enumerate(loocv_with_anchor_indices(cond_keys, always_train_keys), start=1):
        # ----- single inner split: randomly pick k groups for validation -----
        inner_train_idx, inner_valid_idx, valid_groups = make_random_group_valid(
            train_idx, cond_keys, always_train_keys, k=k_valid_groups, rng=rng
        )

        X_train, y_train = X[inner_train_idx], y[inner_train_idx]
        # ===== augment labels on the train partition only =====
        X_train_aug, y_train_aug = augment_labels_normal(
            X_train, y_train, K=K_AUG, rng=rng, min_gr=0.0, truncated=True
        )

        fold_pipe = clone(pipe)

        # Inject validation for EarlyStopping (exactly like ValidSplit but predefined)
        if inner_valid_idx.size > 0:
            X_valid, y_valid = X[inner_valid_idx], y[inner_valid_idx]
            valid_ds = Dataset(X_valid.astype(np.float32, copy=False),
                               y_valid.astype(np.float32, copy=False))
            fold_pipe.set_params(
                net__train_split=predefined_split(valid_ds),
                net__callbacks=[LRScheduler(policy='ReduceLROnPlateau', monitor='valid_loss', factor=0.5, patience=5, min_lr=1e-6, cooldown=0),
                                GradientNormClipping(1.0),
                                EarlyStopping(patience=50, monitor='valid_loss', load_best=True)],
                net__criterion__min_var=1e-4,
                net__iterator_train__shuffle=False,
                net__iterator_train__num_workers=0,
                net__iterator_valid__shuffle=False,
                net__iterator_valid__num_workers=0,
            )
        else:
            fold_pipe.set_params(
                net__train_split=None, net__callbacks=[],
                net__iterator_train__shuffle=False,
                net__iterator_train__num_workers=0,
                net__iterator_valid__shuffle=False,
                net__iterator_valid__num_workers=0,
            )
            
        set_seed(42)
        # >>> minimal but critical change: use augmented data <<<
        fold_pipe.fit(X_train_aug, y_train_aug)

        # ===== outer test with the trained model =====
        X_test, y_test = X[test_idx], y[test_idx]
        y_pred = fold_pipe.predict(X_test).ravel()
        y_true_vec = y_test[:, 0].ravel()   # GR_mean per replicate
        y_sd_vec   = y_test[:, 1].ravel()   # provided true SD per replicate (identical within condition)

        # ---------- replicate-level metrics ----------
        mse_rep  = mean_squared_error(y_true_vec, y_pred)
        rmse_rep = np.sqrt(mse_rep)
        mae_rep  = mean_absolute_error(y_true_vec, y_pred)
        mse_scores.append(mse_rep); rmse_scores.append(rmse_rep)
        mae_scores.append(mae_rep)

        # store replicate predictions
        y_pred_all[test_idx] = y_pred

        # ================= condition-level metrics via per-condition means =================
        conds_in_fold = np.unique(cond_keys[test_idx])
        for group_name in conds_in_fold:
            loc = np.where(cond_keys[test_idx] == group_name)[0]
            y_true_c = y_true_vec[loc]
            y_pred_c = y_pred[loc]
            y_sd_c   = y_sd_vec[loc]

            # per-condition means
            y_true_scalar = float(np.mean(y_true_c))
            y_pred_mean   = float(np.mean(y_pred_c))

            # metrics based on condition means
            cond_mse  = (y_pred_mean - y_true_scalar) ** 2
            cond_rmse = np.sqrt(cond_mse)
            cond_mae  = abs(y_pred_mean - y_true_scalar)
            cond_mse_scores.append(cond_mse)
            cond_rmse_scores.append(cond_rmse)
            cond_mae_scores.append(cond_mae)
            
            # store per-condition means (the fold where this condition is tested provides its estimate)
            cond_pred_dict[group_name] = y_pred_mean
            cond_true_dict[group_name] = y_true_scalar
        # =====================================================================

        # ----- robust inspection (best epoch / valid loss / ES flag) -----
        net = fold_pipe.named_steps['net']
        # print("skorch device:", net.device)   # should say 'cuda'
        epochs_run = len(net.history)
        early_stopped = epochs_run < net.max_epochs

        best_epoch = None
        best_valid_loss = None
        if epochs_run > 0 and 'valid_loss' in net.history[0]:
            try:
                best_row = min(net.history, key=lambda r: r.get('valid_loss', float('inf')))
                best_epoch = best_row.get('epoch')
                best_valid_loss = best_row.get('valid_loss')
            except ValueError:
                pass
        try:
            es = next(cb for name, cb in net.callbacks_ if isinstance(cb, EarlyStopping))
            best_epoch = getattr(es, 'best_epoch_', best_epoch)
            best_valid_loss = getattr(es, 'best_score_', best_valid_loss)
        except StopIteration:
            pass

        fold_info.append({
            "fold": fold_idx,
            "outer_test_idx": int(test_idx[0]),
            "outer_test_cond": str(conds_in_fold[0]),
            "n_train_fold": int(len(train_idx)),
            "n_train_internal": int(len(inner_train_idx)),
            "n_valid_internal": int(len(inner_valid_idx)),
            "valid_groups": list(valid_groups),
            "epochs_run": epochs_run,
            "best_epoch": None if best_epoch is None else int(best_epoch),
            "best_valid_loss": None if best_valid_loss is None else float(best_valid_loss),
            "early_stopped": bool(early_stopped),
        })

        import gc, torch
        del fold_pipe
        gc.collect()
        if torch.cuda.is_available():
            torch.cuda.empty_cache()

    # ---------- replicate-level summary ----------
    replicate_summary = {
        'MSE_mean':  float(np.mean(mse_scores)),  'MSE_sd':  float(np.std(mse_scores)),
        'RMSE_mean': float(np.mean(rmse_scores)), 'RMSE_sd': float(np.std(rmse_scores)),
        'MAE_mean':  float(np.mean(mae_scores)),  'MAE_sd':  float(np.std(mae_scores)),
        'y_pred_all': y_pred_all,
        'fold_info': fold_info,
    }

    # ---------- condition-level summary (+ Spearman on condition means) ----------
    COND_MSE_mean  = float(np.mean(cond_mse_scores));  COND_MSE_sd  = float(np.std(cond_mse_scores))
    COND_RMSE_mean = float(np.mean(cond_rmse_scores)); COND_RMSE_sd = float(np.std(cond_rmse_scores))
    
    cond_names = list(cond_true_dict.keys())
    true_vec   = np.array([cond_true_dict[k] for k in cond_names]) if cond_names else np.array([])
    pred_vec   = np.array([cond_pred_dict[k] for k in cond_names]) if cond_names else np.array([])
    if true_vec.size > 1 and pred_vec.size > 1:
        spearman_rho, spearman_p = spearmanr(true_vec, pred_vec)
        pearson_r, pearson_p = pearsonr(true_vec, pred_vec)
    else:
        spearman_rho, spearman_p = np.nan, np.nan
        pearson_r, pearson_p = np.nan, np.nan

    return {
        **replicate_summary,
        'COND_MSE_mean':  COND_MSE_mean,   'COND_MSE_sd':  COND_MSE_sd,
        'COND_RMSE_mean': COND_RMSE_mean,  'COND_RMSE_sd': COND_RMSE_sd,
        'COND_SPEARMAN_R': float(spearman_rho),
        'COND_SPEARMAN_P': float(spearman_p),
        'COND_PEARSON_R': float(pearson_r),
        'COND_PEARSON_P': float(pearson_p),

        # expose per-condition means and CVs for inspection
        'COND_MEAN_TRUE': cond_true_dict,
        'COND_MEAN_PRED': cond_pred_dict,
    }

def print_summary(name, summary, y, show_condition_level=True):
    """
    Pretty-print LOOCV results (safe ASCII only) with optional condition-level
    metrics and CV diagnostics.
    """
    import numpy as np
    from scipy.stats import pearsonr

    # Use GR_mean if y has two columns; otherwise use y as-is
    y_true = np.asarray(y)
    if y_true.ndim == 2 and y_true.shape[1] >= 1:
        y_true = y_true[:, 0]
    y_true = y_true.ravel()
    y_pred_all = np.asarray(summary.get('y_pred_all', []))

    # Safe Pearson (mask non-finite and size-align)
    mask = np.isfinite(y_true) & np.isfinite(y_pred_all) & (np.arange(y_true.size) < y_pred_all.size)
    if mask.sum() >= 2:
        r, p = pearsonr(y_true[mask], y_pred_all[mask])
    else:
        r, p = np.nan, np.nan

    # ---------------- Replicate-level prints ----------------
    print(f"\n=== {name} (LOOCV) ===")
        # ---------------- Condition-level metrics ----------------
    if show_condition_level and 'COND_MSE_mean' in summary:
        print("\n-- Condition-level (means over replicates) --")
        print(f"RMSE:  {summary.get('COND_RMSE_mean', float('nan')):.6f} +/- {summary.get('COND_RMSE_sd', float('nan')):.6f}")
        if 'COND_SPEARMAN_R' in summary:
            print(f"Spearman (cond means): {summary.get('COND_SPEARMAN_R', float('nan')):.4f} "
                  f"(p={summary.get('COND_SPEARMAN_P', float('nan')):.3g})")
        if 'COND_PEARSON_R' in summary:
            print(
                f"Pearson (cond means):  {summary.get('COND_PEARSON_R', float('nan')):.4f} "
                f"(p={summary.get('COND_PEARSON_P', float('nan')):.3g})"
            )
# ===============================
# Plans (parent-side RNG only)
# ===============================
def split_train_test(cond_keys: np.ndarray, tg: Sequence[Any], always_train_keys: Sequence[Any]):
    test_mask = np.isin(cond_keys, tg)
    test_idx = np.where(test_mask)[0]
    train_idx = np.where(~test_mask)[0]
    if np.isin(np.where(np.isin(cond_keys, list(always_train_keys)))[0], test_idx).any():
        raise RuntimeError("Anchor condition leaked into test.")
    return train_idx, test_idx

def make_iter_plan(
    X: np.ndarray,
    cond_keys: np.ndarray,
    always_train_keys: Sequence[Any],
    iter_groups: Sequence[Sequence[Any]],
    *,
    k_valid_groups: int,
    rng,  # int or RandomState
    bg_size: int,
    explain_max: int,
):
    # --- minimal, robust change: normalize RNG ---
    rs = rng if hasattr(rng, "choice") else np.random.RandomState(int(rng))
    plans = []
    for it, tg in enumerate(iter_groups):
        tg = tuple(tg)
        seed_it = 42

        train_idx, test_idx = split_train_test(cond_keys, tg, always_train_keys)

        inner_train_idx, inner_valid_idx, _ = make_random_group_valid(
            train_idx,  cond_keys, always_train_keys, k=k_valid_groups, rng=seed_it
        )

        if len(train_idx) > bg_size:
            local_bg = rs.choice(len(train_idx), size=bg_size, replace=False)
            bg_idx = train_idx[local_bg]
        else:
            bg_idx = train_idx.copy()

        if len(test_idx) > explain_max:
            local_te = rs.choice(len(test_idx), size=explain_max, replace=False)
            te_sub_idx = test_idx[local_te]
        else:
            te_sub_idx = test_idx.copy()

        plans.append(dict(
            it=int(it),
            seed=seed_it,
            test_groups=list(tg),
            train_idx=inner_train_idx,
            inner_valid_idx=inner_valid_idx,
            test_idx=test_idx,
            bg_idx=bg_idx,
            te_sub_idx=te_sub_idx,
        ))
    return plans

# ===============================
# Worker (top-level, no closures)
# ===============================
def compute_one_iter(
    *,
    pipe_factory_name: str,
    plan: Dict[str, Any],
    X: np.ndarray,
    y: np.ndarray,
    cond_keys: np.ndarray,
    always_train_keys: Sequence[Any],
    K_AUG: int,
    nsamples: int | str,
    use_cuda: bool,
    device_ids: Optional[Sequence[int]],
    train_batch_size: Optional[int],
    use_cuda_deterministic: bool,
    compute_anchor_groups: bool,
) -> Dict[str, Any]:
    # ------------ imports & warning filters (apply once per worker) ------------
    import math
    import random
    import re
    import warnings
    import numpy as np
    from collections import defaultdict
    from sklearn.metrics import mean_squared_error

    # IMPORTANT: message must be a STRING regex, not a compiled re.Pattern
    warnings.filterwarnings(
        "ignore",
        message=r"Changing solver from 'eig' to 'svd'",
        category=UserWarning,
    )
    # Optional noise guards (safe):
    warnings.filterwarnings("ignore", category=FutureWarning)
    warnings.filterwarnings("ignore", message=r".*overflow.*", category=RuntimeWarning)


    # ------------ setup / device pinning ------------
    it   = int(plan["it"])
    seed = int(plan["seed"])
    set_seed(42)  # keep determinism parity

    template_pipe = pipe_factory_name()
    pipe = clone(template_pipe)

    if use_cuda and device_ids:
        dev_id = device_ids[it % len(device_ids)]
        try:
            pipe.set_params(net__device=f"cuda:{dev_id}")
            if train_batch_size is not None:
                pipe.set_params(net__iterator_train__batch_size=int(train_batch_size))
        except Exception:
            pass
    else:
        try:
            pipe.set_params(net__device="cpu")
        except Exception:
            pass

    train_idx = plan["train_idx"]
    valid_idx = plan["inner_valid_idx"]
    test_idx  = plan["test_idx"]
    bg_idx    = plan["bg_idx"]
    te_idx    = plan["te_sub_idx"]

    # ------------ data splits ------------
    X_train, y_train = X[train_idx], y[train_idx]
    if (K_AUG is not None) and (K_AUG > 1):
        X_train, y_train = augment_labels_normal(
            X_train, y_train, K=K_AUG, rng=seed, min_gr=0.0, truncated=True
        )

    if valid_idx is not None and len(valid_idx) > 0:
        X_valid, y_valid = X[valid_idx], y[valid_idx]
        valid_ds = Dataset(
            X_valid.astype(np.float32, copy=False),
            y_valid.astype(np.float32, copy=False),
        )
        try:
            pipe.set_params(
                net__train_split=predefined_split(valid_ds),
                net__callbacks=[
                    LRScheduler(policy="ReduceLROnPlateau", monitor="valid_loss",
                                factor=0.5, patience=5, min_lr=1e-6, cooldown=0),
                    GradientNormClipping(1.0),
                    EarlyStopping(patience=50, monitor="valid_loss", load_best=True),
                ],
                net__criterion__min_var=1e-4,
                net__verbose=0,
                net__iterator_train__shuffle=False,
                net__iterator_train__num_workers=0,
                net__iterator_valid__shuffle=False,
                net__iterator_valid__num_workers=0,
            )
        except Exception:
            pass
    else:
        try:
            pipe.set_params(
                net__train_split=None,
                net__callbacks=[],
                net__verbose=0,
                net__iterator_train__shuffle=False,
                net__iterator_train__num_workers=0,
                net__iterator_valid__shuffle=False,
                net__iterator_valid__num_workers=0,
            )
        except Exception:
            pass

    # # ------------ fit / baseline RMSE ------------
    # set_seed(42)
    # pipe.fit(X_train, y_train)

    # X_test_full = X[test_idx]
    # y_test_full = np.asarray(y[test_idx][:, 0]).reshape(-1)  # robust to 1D/2D y
    # y_pred_full = pipe.predict(X_test_full).ravel()
    # base_rmse = math.sqrt(mean_squared_error(y_test_full, y_pred_full))

    # # ------------ SHAP background & targets ------------
    # bg = np.ascontiguousarray(np.asarray(X[bg_idx], dtype=np.float32)).reshape(-1, X.shape[1])
    # X_te_explain = np.ascontiguousarray(np.asarray(X[te_idx], dtype=np.float32)).reshape(-1, X.shape[1])
    # te_sub_global_idx = te_idx

    # # ===== cuML Kernel SHAP (GPU) with NumPy<->CuPy bridge =====
    # import cupy as cp

    # bg_cp   = cp.asarray(bg)
    # X_te_cp = cp.asarray(X_te_explain)

    # def f_predict(Z):
        # # cuML passes CuPy; pipeline wants NumPy; return CuPy back.
        # Z_np = cp.asnumpy(Z) if isinstance(Z, cp.ndarray) else np.asarray(Z, dtype=np.float32)
        # y_np = pipe.predict(Z_np).ravel().astype(np.float32, copy=False)
        # return cp.asarray(y_np)

    # use_cuml_gpu=False
    # if use_cuml_gpu:
        # try:    
            # from cuml.explainer import KernelExplainer as _CuMLKernelExplainer
            # use_cuml_gpu = True
            # np.random.seed(42); random.seed(42)
            # explainer = _CuMLKernelExplainer(
                # model=f_predict,
                # data=bg_cp,
                # nsamples=nsamples,
                # random_state=42,
                # is_gpu_model=True,  # callback returns CuPy
            # )
            # np.random.seed(42); random.seed(42)
            # shap_vals = explainer.shap_values(X_te_cp)
            # if it==0: 
                # print("Used GPU-based SHAP KernelExplainer.")
        # except Exception:
            # use_cuml_gpu = False

    # # Fallback to CPU SHAP if cuML not available or failed
    # if not use_cuml_gpu:
        # import shap, random
        # np.random.seed(42); random.seed(42)
        # f_predict_cpu = lambda Z: pipe.predict(Z).ravel()
        # explainer = shap.KernelExplainer(f_predict_cpu, X[bg_idx], seed=42)
        # np.random.seed(42); random.seed(42)
        # shap_vals = explainer.shap_values(X_te_explain, nsamples=nsamples, silent=True)
        # if it==0: 
            # print("Used CPU-based SHAP KernelExplainer.")

    # # Normalize list and move to NumPy for downstream ops
    # if isinstance(shap_vals, list):
        # shap_vals = shap_vals[0]
    # try:
        # shap_vals = cp.asnumpy(shap_vals)
    # except Exception:
        # pass
        
    # ------------ fit / baseline RMSE ------------
    set_seed(42)
    pipe.fit(X_train, y_train)

    X_test_full = X[test_idx]
    y_test_full = np.asarray(y[test_idx][:, 0]).reshape(-1)  # robust to 1D/2D y
    y_pred_full = pipe.predict(X_test_full).ravel()
    base_rmse = math.sqrt(mean_squared_error(y_test_full, y_pred_full))

    # ------------ SHAP background & targets ------------
    bg = np.ascontiguousarray(np.asarray(X[bg_idx], dtype=np.float32)).reshape(-1, X.shape[1])
    X_te_explain = np.ascontiguousarray(np.asarray(X[te_idx], dtype=np.float32)).reshape(-1, X.shape[1])
    te_sub_global_idx = te_idx

    # ===== Try cuML Kernel SHAP (GPU) with NumPy<->CuPy bridge =====
    use_cuml_gpu = False

    try:
        import cupy as cp
        from cuml.explainer import KernelExplainer as _CuMLKernelExplainer

        use_cuml_gpu = True

        bg_cp   = cp.asarray(bg)
        X_te_cp = cp.asarray(X_te_explain)

        def f_predict(Z):
            # cuML passes CuPy; pipeline wants NumPy; return CuPy back.
            Z_np = cp.asnumpy(Z) if isinstance(Z, cp.ndarray) else np.asarray(Z, dtype=np.float32)
            y_np = pipe.predict(Z_np).ravel().astype(np.float32, copy=False)
            return cp.asarray(y_np)

        np.random.seed(42); random.seed(42)
        explainer = _CuMLKernelExplainer(
            model=f_predict,
            data=bg_cp,
            nsamples=nsamples,
            random_state=42,
            is_gpu_model=True,  # callback returns CuPy
        )
        np.random.seed(42); random.seed(42)
        shap_vals = explainer.shap_values(X_te_cp)
        if it == 0:
            print("Used GPU-based SHAP KernelExplainer.")

    except Exception:
        # Anything fails (no cupy, no cuml, GPU error, etc.) → fall back to CPU
        use_cuml_gpu = False

    # Fallback to CPU SHAP if cuML not available or failed
    if not use_cuml_gpu:
        import shap, random
        np.random.seed(42); random.seed(42)
        f_predict_cpu = lambda Z: pipe.predict(Z).ravel()
        explainer = shap.KernelExplainer(f_predict_cpu, X[bg_idx], seed=42)
        np.random.seed(42); random.seed(42)
        shap_vals = explainer.shap_values(X_te_explain, nsamples=nsamples, silent=True)
        if it == 0:
            print("Used CPU-based SHAP KernelExplainer.")

    # Normalize list and move to NumPy for downstream ops
    if isinstance(shap_vals, list):
        shap_vals = shap_vals[0]

    # Safely try to convert from CuPy to NumPy if CuPy exists
    try:
        import cupy as cp
        if isinstance(shap_vals, cp.ndarray):
            shap_vals = cp.asnumpy(shap_vals)
    except Exception:
        pass  # shap_vals is already a NumPy array (CPU path)

    # ------------ SHAP aggregation (main targets) ------------
    abs_vals       = np.abs(shap_vals)
    local_abs_sum  = abs_vals.sum(axis=0)
    local_abs_sumsq= (abs_vals ** 2).sum(axis=0)
    local_n_rows   = int(abs_vals.shape[0])

    by_group_sum_loc = defaultdict(lambda: np.zeros(abs_vals.shape[1], dtype=np.float64))
    by_group_n_loc   = defaultdict(int)

    these_groups = cond_keys[te_sub_global_idx]
    for g in np.unique(these_groups):
        m = (these_groups == g)
        by_group_sum_loc[g] += abs_vals[m].sum(axis=0)
        by_group_n_loc[g]   += int(m.sum())

    # ------------ ANCHOR groups (GPU-consistent) ------------
    if compute_anchor_groups and len(always_train_keys) > 0:
        anchor_mask_all = np.isin(cond_keys, list(always_train_keys))
        anchor_idx_all  = np.where(anchor_mask_all)[0]
        anchor_idx      = np.intersect1d(train_idx, anchor_idx_all, assume_unique=False)
        if anchor_idx.size > 0:
            X_anchor_explain = np.ascontiguousarray(
                X[anchor_idx].astype(np.float32, copy=False)
            )
            if use_cuml_gpu:
                X_anchor_cp = cp.asarray(X_anchor_explain)
                shap_vals_anc = explainer.shap_values(X_anchor_cp) # <-- CuPy in
                if isinstance(shap_vals_anc, list):
                    shap_vals_anc = shap_vals_anc[0]
                try:
                    shap_vals_anc = cp.asnumpy(shap_vals_anc)                             # --> NumPy out
                except Exception:
                    pass
            else:
                import shap
                shap_vals_anc = explainer.shap_values(X_anchor_explain, nsamples=nsamples, silent=True)
                if isinstance(shap_vals_anc, list):
                    shap_vals_anc = shap_vals_anc[0]

            abs_vals_anc = np.abs(shap_vals_anc)
            these_groups_anc = cond_keys[anchor_idx]
            for g in np.unique(these_groups_anc):
                m = (these_groups_anc == g)
                key = f"ANCHOR:{g}"
                by_group_sum_loc[key] += abs_vals_anc[m].sum(axis=0)
                by_group_n_loc[key]   += int(m.sum())

    # ------------ cleanup ------------
    import gc, torch
    del pipe
    gc.collect()
    if torch.cuda.is_available():
        torch.cuda.empty_cache()

    # ------------ debug payload for parent (since workers don't print) ------------
    debug = {
        "it": it,
        "X_train_shape": tuple(np.asarray(X_train).shape),
        "y_train_shape": tuple(np.asarray(y_train).shape),
        "X_test_full_shape": tuple(np.asarray(X_test_full).shape),
        "y_test_full_shape": tuple(np.asarray(y_test_full).shape),
        "bg_shape": tuple(np.asarray(bg).shape),
        "X_te_explain_shape": tuple(np.asarray(X_te_explain).shape),
        "y_pred_full_shape": tuple(np.asarray(y_pred_full).shape),
        "shap_vals_shape": tuple(np.asarray(shap_vals).shape) if shap_vals is not None else None,
        "nsamples": nsamples,
        "use_cuml_gpu": use_cuml_gpu,
    }

    return {
        "it": it,
        "test_groups": plan["test_groups"],
        "baseline_rmse": float(base_rmse),
        "abs_sum": local_abs_sum,
        "abs_sumsq": local_abs_sumsq,
        "n_rows": local_n_rows,
        "by_group_sum": by_group_sum_loc,
        "by_group_n": by_group_n_loc,
        "debug": debug,
    }


# ---- NEW TOP LEVEL: safe dispatcher (no closures) ----
def safe_dispatch(pipe_factory_name, plan, X, y, cond_keys, always_train_keys,
                K_AUG, nsamples, use_cuda, device_ids, train_batch_size,
                use_cuda_deterministic, compute_anchor_groups):
    
    return compute_one_iter(
            pipe_factory_name=pipe_factory_name,   # <-- pass it
            plan=plan,
            X=X, y=y, cond_keys=cond_keys, always_train_keys=always_train_keys,
            K_AUG=K_AUG, nsamples=nsamples,
            use_cuda=use_cuda, device_ids=device_ids, train_batch_size=train_batch_size,
            use_cuda_deterministic=use_cuda_deterministic,
            compute_anchor_groups=compute_anchor_groups,
        )
    # try:
        # return compute_one_iter(
            # pipe_factory_name=pipe_factory_name,   # <-- pass it
            # plan=plan,
            # X=X, y=y, cond_keys=cond_keys, always_train_keys=always_train_keys,
            # K_AUG=K_AUG, nsamples=nsamples,
            # use_cuda=use_cuda, device_ids=device_ids, train_batch_size=train_batch_size,
            # use_cuda_deterministic=use_cuda_deterministic,
            # compute_anchor_groups=compute_anchor_groups,
        # )
    # except Exception as e:
        # return ("__ERROR__", {"iter": plan.get("it", -1),
                            # "groups": plan.get("test_groups", []),
                            # "error": repr(e)})


# ===============================
# ===============================
def shap_importance_mc_deterministic(
    *,
    pipe_factory_name: str,
    X: np.ndarray,
    y: np.ndarray,
    cond_keys: np.ndarray,
    always_train_keys: Sequence[Any],
    feature_names: Optional[Sequence[str]] = None,
    n_iters: int = 200,
    test_k: int = 3,
    exhaustive: bool = True,
    shuffle_combos: bool = False,
    k_valid_groups: int = 3,
    K_AUG: int = 100,
    bg_size: int = 200,           # currently not effective
    explain_max: int = 400,       # currently not effective as well
    nsamples: int | str = 1024,   # fixed budget for parity; can use "auto"
    compute_anchor_groups: bool = True,
    n_jobs: int = 1,
    backend: str = "loky",
    prefer: str = "processes",
    rng_seed: int = 42,
    use_cuda: bool = False,
    device_ids: Optional[Sequence[int]] = None,
    train_batch_size: Optional[int] = None,
    use_cuda_deterministic: bool = False,
):


    X = np.asarray(X); y = np.asarray(y); cond_keys = np.asarray(cond_keys)
    d = X.shape[1]
    if feature_names is None:
        feature_names = [f"feat_{j}" for j in range(d)]

    all_groups = np.unique(cond_keys)
    anchor_set = set(always_train_keys)
    non_anchor_groups = np.array([g for g in all_groups if g not in anchor_set], dtype=object)
    if len(non_anchor_groups) < test_k:
        raise ValueError(f"Not enough non-anchor groups for test_k={test_k}.")

    from itertools import combinations
    rs = np.random.RandomState(rng_seed)
    if exhaustive:
        combo_list = [tuple(c) for c in combinations(non_anchor_groups, test_k)]
        if shuffle_combos:
            rs.shuffle(combo_list)
        iter_groups = combo_list
    else:
        iter_groups = [tuple(rs.choice(non_anchor_groups, size=test_k, replace=False))
                       for _ in range(n_iters)]

    plans = make_iter_plan(
        X=X, cond_keys=cond_keys, always_train_keys=always_train_keys,
        iter_groups=iter_groups, k_valid_groups=k_valid_groups,
        rng=42, bg_size=bg_size, explain_max=explain_max
    )
    
    # par_out = safe_dispatch(pipe_factory_name=pipe_factory_name, plan=plans[0],
                # X=X, y=y, cond_keys=cond_keys, always_train_keys=always_train_keys,
                # K_AUG=K_AUG, nsamples=nsamples,
                # use_cuda=use_cuda, device_ids=device_ids,
                # train_batch_size=None, use_cuda_deterministic=use_cuda_deterministic,
                # compute_anchor_groups=compute_anchor_groups,
                # ) 
    # --- progress bar (tqdm_joblib if available; else joblib verbose) ---
    from tqdm.auto import tqdm
    from tqdm_joblib import tqdm_joblib
    from joblib import Parallel, delayed, parallel_backend
    import warnings
    
    # optional: quiet down sklearn's LARS/Lasso messages
    from sklearn.exceptions import ConvergenceWarning
    warnings.filterwarnings("ignore", category=ConvergenceWarning)

    def _pick_n_jobs(min_jobs=2, max_jobs=5, target_vram_frac=0.80):
        """
        Choose n_jobs based on current GPU VRAM usage so workers don't fight for memory.
        Falls back to max_jobs if GPU info isn't available.
        """
        try:
            import cupy as cp
            free, total = cp.cuda.runtime.memGetInfo()   # bytes
            used_frac = 1.0 - (free / total)
            # how much headroom remains vs target
            headroom = max(0.0, target_vram_frac - used_frac) / max(target_vram_frac, 1e-6)
            n = int(min_jobs + headroom * (max_jobs - min_jobs))
            return max(min_jobs, min(max_jobs, n))
        except Exception:
            return max_jobs
    n_jobs_ = _pick_n_jobs(min_jobs=1, max_jobs=n_jobs)
    with tqdm_joblib(tqdm(total=len(plans), desc="LOO Iterations", leave=True, dynamic_ncols=True)):
        with parallel_backend('loky', timeout=3600):
            par_out = Parallel(n_jobs=n_jobs, backend=backend, prefer=prefer, verbose=10)(
                delayed(safe_dispatch)(
                    pipe_factory_name=pipe_factory_name, plan=plan,
                    X=X, y=y, cond_keys=cond_keys, always_train_keys=always_train_keys,
                    K_AUG=K_AUG, nsamples=nsamples,
                    use_cuda=use_cuda, device_ids=device_ids,
                    train_batch_size=None, use_cuda_deterministic=use_cuda_deterministic,
                    compute_anchor_groups=compute_anchor_groups,
                    ) for plan in plans
            )
        
        # # DEBUG ++++++++++++++++++++++++++++++++++++++++++
        # for item in par_out:
            # if isinstance(item, tuple) and item[0] == "__ERROR__":
                # continue
            # d = item.get("debug", {})
            # print(
                # f"[it={d.get('it')}] "
                # f"X_train{d.get('X_train_shape')} y_train{d.get('y_train_shape')} | "
                # f"test{d.get('X_test_full_shape')} y_test{d.get('y_test_full_shape')} | "
                # f"bg{d.get('bg_shape')} te{d.get('X_te_explain_shape')} | "
                # f"pred{d.get('y_pred_full_shape')} shap{d.get('shap_vals_shape')} | "
                # f"nsamples={d.get('nsamples')} backend={'cuML_GPU' if d.get('use_cuml_gpu') else 'CPU'}"
            # )
        # # DEBUG ++++++++++++++++++++++++++++++++++++++++++

    abs_shap_sums  = np.zeros(d, dtype=np.float64)
    abs_shap_sumsq = np.zeros(d, dtype=np.float64)
    n_rows_total   = 0
    baseline_rmse_list = []
    test_groups_log = []
    by_group_sum = defaultdict(lambda: np.zeros(d, dtype=np.float64))
    by_group_n   = defaultdict(int)
    errors = []

    for item in par_out:
        if isinstance(item, tuple) and item[0] == "__ERROR__":
            errors.append(item[1]); continue
        r = item
        abs_shap_sums  += r["abs_sum"]
        abs_shap_sumsq += r["abs_sumsq"]
        n_rows_total   += r["n_rows"]
        baseline_rmse_list.append(r["baseline_rmse"])
        test_groups_log.append(r["test_groups"])
        for g, v in r["by_group_sum"].items():
            by_group_sum[g] += v
        for g, n in r["by_group_n"].items():
            by_group_n[g] += n

    by_group_df = None
    if len(by_group_n) > 0:
        cols, mats = [], []
        for g, n_g in by_group_n.items():
            if n_g > 0:
                mats.append(by_group_sum[g] / float(n_g))
                cols.append(g)
        if mats:
            M = np.vstack(mats).T
            by_group_df = pd.DataFrame(M, index=feature_names, columns=[str(c) for c in cols])
            by_group_df.index.name = "feature"

    baseline_rmse_arr = np.array(baseline_rmse_list, dtype=np.float32)
    mean_abs_shap = abs_shap_sums / max(1, n_rows_total)
    var_abs_shap  = (abs_shap_sumsq / max(1, n_rows_total)) - mean_abs_shap**2
    var_abs_shap  = np.maximum(var_abs_shap, 0.0)
    sd_abs_shap   = np.sqrt(var_abs_shap)

    mean_base_rmse = float(np.mean(baseline_rmse_arr)) if baseline_rmse_arr.size else 1.0
    norm_mean_abs  = mean_abs_shap / (mean_base_rmse + 1e-12)

    n_eff = float(max(1, n_rows_total))
    se = sd_abs_shap / np.sqrt(n_eff)
    z = 1.96
    ci_lo = mean_abs_shap - z * se
    ci_hi = mean_abs_shap + z * se

    shap_df = (
        pd.DataFrame({
            "feature": feature_names,
            "mean_abs_shap": mean_abs_shap,
            "sd_abs_shap": sd_abs_shap,
            "ci95_lo": ci_lo,
            "ci95_hi": ci_hi,
            "mean_abs_shap_over_baseRMSE": norm_mean_abs,
            "n_explained_rows": int(n_rows_total),
            "n_iters": int(len(baseline_rmse_arr)),
            "bg_size": int(min(bg_size, len(X)))
        })
        .sort_values("mean_abs_shap", ascending=False)
        .reset_index(drop=True)
    )

    
    raw = {
        "baseline_rmse": baseline_rmse_arr,
        "test_groups_log": test_groups_log,
        "n_iters_eff": int(len(baseline_rmse_arr)),
        "n_rows_total": n_rows_total,
        "by_group_df": by_group_df,
        "errors": errors,
        "plans_used": plans,
    }
    return shap_df, raw


# ===============================
# ===============================

def _ap_pref_from_quantile(S, q=0.5):
    """
    Preference scalar for AP from a similarity matrix S by upper-tri quantile.
    Larger preference -> more clusters.
    """
    tri = S[np.triu_indices_from(S, k=1)]
    if tri.size == 0:
        return np.median(S)
    return float(np.quantile(tri, q))

def ap_cluster_features_global(
    by_group_df: pd.DataFrame,
    scale=True,
    similarity="cosine",     # 'cosine' (precomputed) or 'euclidean' (AP's built-in -||x-y||^2)
    pref_quantile=0.50,      # median by default
    damping=0.80,
    random_state=42
):
    """
    Global AP clustering of FEATURES using per-condition mean|SHAP| signatures (rows = features).
    Returns:
      assign (pd.Series: feature -> cluster_id),
      exemplars (pd.Series: cluster_id -> exemplar feature),
      model (sklearn AffinityPropagation)
    """
    if by_group_df is None or by_group_df.empty:
        raise ValueError("by_group_df is empty. Enable return_by_group in SHAP step.")
    X = by_group_df.values  # (d x G)

    if scale:
        X = StandardScaler().fit_transform(X)

    if similarity == "cosine":
        S = cosine_similarity(X)  # higher = more similar
        pref = _ap_pref_from_quantile(S, pref_quantile)
        ap = AffinityPropagation(affinity="precomputed", damping=damping,
                                 preference=pref, random_state=random_state)
        ap.fit(S)
        labels = ap.labels_
        centers = ap.cluster_centers_indices_
    elif similarity == "euclidean":
        # AP will internally use -||x - y||^2; we still pick a preference scalar
        # from implied similarities to control cluster count.
        D2 = pairwise_distances(X, metric="euclidean") ** 2
        S = -D2
        pref = _ap_pref_from_quantile(S, pref_quantile)
        ap = AffinityPropagation(affinity="euclidean", damping=damping,
                                 preference=pref, random_state=random_state)
        ap.fit(X)
        labels = ap.labels_
        centers = ap.cluster_centers_indices_
    else:
        raise ValueError("similarity must be 'cosine' or 'euclidean'.")

    # Build outputs
    feats = by_group_df.index.to_list()
    assign = pd.Series(labels, index=feats, name="cluster")
    # exemplar feature names
    if centers is None:
        # rare non-convergence: fallback to one exemplar per cluster by max diag(S)
        # (still deterministic given random_state/scale)
        centers = []
        for c in np.unique(labels):
            idx = np.where(labels == c)[0]
            if similarity == "cosine":
                # pick the feature with max average similarity to its cluster
                local = S[np.ix_(idx, idx)]
                centers.append(idx[np.argmax(local.mean(axis=1))])
            else:
                local = S[np.ix_(idx, idx)]  # S is -D^2
                centers.append(idx[np.argmax(local.mean(axis=1))])
        centers = np.array(centers, dtype=int)
    exemplar_idx = {int(c): int(i) for c, i in zip(np.unique(labels), centers)}
    exemplars = pd.Series({c: by_group_df.index[i] for c, i in exemplar_idx.items()}, name="exemplar")

    return assign, exemplars, ap

def ap_refine_local_subclusters(
    by_group_df: pd.DataFrame,
    global_assign: pd.Series,
    scale=True,
    similarity="cosine",
    pref_quantile=0.50,
    damping=0.80,
    random_state=42,
    min_cluster_size=6
):
    """
    Run AP *within each global cluster* to find local substructure.
    Returns a DataFrame with columns: feature, parent_cluster, subcluster (local id), local_exemplar.
    Skips tiny clusters (< min_cluster_size).
    """
    rows = []
    for c in np.unique(global_assign.values):
        members = global_assign.index[global_assign.values == c]
        if len(members) < min_cluster_size:
            for f in members:
                rows.append({"feature": f, "parent_cluster": int(c), "subcluster": -1, "local_exemplar": f})
            continue
        sub_df = by_group_df.loc[members]
        sub_assign, sub_exemplars, _ = ap_cluster_features_global(
            sub_df, scale=scale, similarity=similarity,
            pref_quantile=pref_quantile, damping=damping, random_state=random_state
        )
        # remap subcluster ids to stable 0..k-1 per parent
        uniq = np.unique(sub_assign.values)
        remap = {old: i for i, old in enumerate(sorted(uniq))}
        local_exemplars = {remap[k]: v for k, v in sub_exemplars.to_dict().items()}
        for f, sc in sub_assign.items():
            rows.append({
                "feature": f,
                "parent_cluster": int(c),
                "subcluster": int(remap[sc]),
                "local_exemplar": local_exemplars[int(remap[sc])]
            })
    return pd.DataFrame(rows).sort_values(["parent_cluster", "subcluster", "feature"]).reset_index(drop=True)

def ap_cluster_conditions(
    by_group_df: pd.DataFrame,
    scale=True,
    similarity="cosine",
    pref_quantile=0.50,
    damping=0.80,
    random_state=42
):
    """
    AP clustering of CONDITIONS (transpose of by_group_df).
    Returns:
      assign_cond (pd.Series: condition -> cluster_id),
      exemplars_cond (pd.Series: cluster_id -> exemplar condition),
      model
    """
    if by_group_df is None or by_group_df.empty:
        raise ValueError("by_group_df is empty.")
    Xc = by_group_df.T.values  # (G x d_features)
    conds = by_group_df.columns.to_list()

    if scale:
        Xc = StandardScaler().fit_transform(Xc)

    if similarity == "cosine":
        S = cosine_similarity(Xc)
        pref = _ap_pref_from_quantile(S, pref_quantile)
        ap = AffinityPropagation(affinity="precomputed", damping=damping,
                                 preference=pref, random_state=random_state)
        ap.fit(S)
        labels, centers = ap.labels_, ap.cluster_centers_indices_
    elif similarity == "euclidean":
        D2 = pairwise_distances(Xc, metric="euclidean") ** 2
        S = -D2
        pref = _ap_pref_from_quantile(S, pref_quantile)
        ap = AffinityPropagation(affinity="euclidean", damping=damping,
                                 preference=pref, random_state=random_state)
        ap.fit(Xc)
        labels, centers = ap.labels_, ap.cluster_centers_indices_
    else:
        raise ValueError("similarity must be 'cosine' or 'euclidean'.")

    assign_cond = pd.Series(labels, index=conds, name="cond_cluster")
    if centers is None:
        centers = []
        for c in np.unique(labels):
            idx = np.where(labels == c)[0]
            local = S[np.ix_(idx, idx)]
            centers.append(idx[np.argmax(local.mean(axis=1))])
        centers = np.array(centers, dtype=int)
    exemplar_idx = {int(c): int(i) for c, i in zip(np.unique(labels), centers)}
    exemplars_cond = pd.Series({c: by_group_df.columns[i] for c, i in exemplar_idx.items()}, name="cond_exemplar")

    return assign_cond, exemplars_cond, ap


# =======================================
def cluster_permutation_importance(
    pipe_template,
    X, y, cond_keys, always_train_keys,
    cluster_assign,                 # pd.Series: index=feature names in model order; values=int cluster id
    n_iters=300, test_k=3, k_valid_groups=3,
    K_AUG=100, rng=42, verbose=False,
    exhaustive=True, shuffle_combos=True,
    conditional=True,               # <-- GLM-residualized conditional permutation
    glm_alpha=1.0,                  # <-- Ridge strength for GLM
):
    """
    Monte Carlo group-CV *cluster*-level permutation importance.
    Measures DELTA-RMSE when disrupting an entire cluster on the held-out test set.

    If conditional=True (recommended):
      - Fit a multioutput Ridge GLM on TRAIN:  X_cluster ~ X_not_cluster
      - On TEST, compute residuals R = X_cluster - E[X_cluster|X_not_cluster]
      - Permute residual ROWS within each held-out condition (keeps residual covariance)
      - Recompose: X_cluster* = E + permuted(R)
      - Predict and compute DELTA-RMSE.

    If conditional=False:
      - Classic within-condition permutation of each column in the cluster (independently).

    Notes:
      - No leakage: GLM fit uses TRAIN only.
      - If a cluster spans all columns (no predictors remain), falls back to unconditional permutation.
      - If a held-out condition has only one test row, no permutation is applied for that group (no effect).
    """
    import numpy as np
    from math import sqrt
    from sklearn.base import clone
    from sklearn.metrics import mean_squared_error
    from tqdm import tqdm

    rs = np.random.RandomState(rng)
    X = np.asarray(X); y = np.asarray(y)
    N, d = X.shape

    # map columns -> cluster ids (assumes cluster_assign.index matches model column order)
    # --- Force cluster_assign shape & dtype, positionally aligned ---
    if isinstance(cluster_assign, pd.Series):
        if len(cluster_assign) != d:
            raise ValueError(f"cluster_assign length {len(cluster_assign)} != X.shape[1] {d}")
        col_clusters = np.asarray(cluster_assign.values, dtype=np.intp)  # positional
    else:
        col_clusters = np.asarray(cluster_assign, dtype=np.intp)
        if col_clusters.shape[0] != d:
            raise ValueError("cluster_assign must have one entry per column in X (positional).")

    clusters = np.unique(col_clusters).astype(np.intp)

    # build iter_groups
    all_groups = np.unique(cond_keys)
    non_anchor_groups = np.array([g for g in all_groups if g not in set(always_train_keys)])

    from itertools import combinations
    if exhaustive:
        combo_list = [tuple(c) for c in combinations(non_anchor_groups, test_k)]
        if shuffle_combos: rs.shuffle(combo_list)
        iter_groups = combo_list
    else:
        iter_groups = [tuple(rs.choice(non_anchor_groups, size=test_k, replace=False)) for _ in range(n_iters)]
    n_iters_eff = len(iter_groups)

    # storage
    dRMSE_sum   = {c: 0.0 for c in clusters}
    dRMSE_sumsq = {c: 0.0 for c in clusters}
    counts      = {c: 0   for c in clusters}
    base_list   = np.zeros(n_iters_eff, dtype=float)

    # helpers from your stack
    from skorch.helper import predefined_split
    from skorch.dataset import Dataset
    from skorch.callbacks import EarlyStopping, LRScheduler, GradientNormClipping

    # GLM
    from sklearn.linear_model import Ridge
    pipe_template = pipe_template()


    pbar = tqdm(iter_groups, total=n_iters_eff, desc="MC cluster DELTA-RMSE", disable=not verbose)
    for it, test_groups in enumerate(pbar):
        test_mask = np.isin(cond_keys, test_groups)
        test_idx  = np.where(test_mask)[0]
        train_idx = np.where(~test_mask)[0]
        # safety: anchors never in test
        if np.isin(np.where(np.isin(cond_keys, list(always_train_keys)))[0], test_idx).any():
            raise RuntimeError("Anchor condition leaked into test.")
        # inner valid split (grouped)
        inner_train_idx, inner_valid_idx, _ = make_random_group_valid(
            train_idx, cond_keys, always_train_keys, k=k_valid_groups, rng=42
        )
        # fit fold
        fold_pipe = clone(pipe_template)
        X_train, y_train = X[inner_train_idx], y[inner_train_idx]
        if K_AUG and K_AUG > 1:
            X_train, y_train = augment_labels_normal(
                X_train, y_train, K=K_AUG, rng=42, min_gr=0.0, truncated=True
            )
            fit_X, fit_y = X_train, y_train
        else:
            fit_X, fit_y = X_train, y_train

        if inner_valid_idx.size > 0:
            X_valid, y_valid = X[inner_valid_idx], y[inner_valid_idx]
            valid_ds = Dataset(X_valid.astype(np.float32, copy=False), y_valid.astype(np.float32, copy=False))
            fold_pipe.set_params(
                net__device='cuda' if torch.cuda.is_available() else 'cpu',                      # use GPU
                net__train_split=predefined_split(valid_ds),
                net__callbacks=[
                    LRScheduler(policy='ReduceLROnPlateau', monitor='valid_loss',
                                factor=0.5, patience=5, min_lr=1e-6, cooldown=0),
                    GradientNormClipping(1.0),
                    EarlyStopping(patience=50, monitor='valid_loss', load_best=True),
                ],
                net__criterion__min_var=1e-4,
                net__verbose=0,
                net__iterator_train__shuffle=False,
                net__iterator_train__num_workers=0,
                net__iterator_valid__shuffle=False,
                net__iterator_valid__num_workers=0,
            )
        else:
            fold_pipe.set_params(net__train_split=None, net__callbacks=[], net__verbose=0,
                    net__device='cuda' if torch.cuda.is_available() else 'cpu',                      # use GPU
                    net__iterator_train__shuffle=False,
                    net__iterator_train__num_workers=0,
                    net__iterator_valid__shuffle=False,
                    net__iterator_valid__num_workers=0,)

        set_seed(42)
        fold_pipe.fit(fit_X, fit_y)

        # baseline
        X_test = X[test_idx]
        y_test = y[test_idx]
        if y_test.ndim > 1: y_test = y_test[:, 0]
        y_test = y_test.ravel()
        y_pred = np.asarray(fold_pipe.predict(X_test)).ravel()
        base_rmse = sqrt(mean_squared_error(y_test, y_pred))
        base_list[it] = base_rmse

        # split test rows by held-out condition for within-group shuffles
        test_groups_vec = cond_keys[test_idx]
        uniq_g = np.unique(test_groups_vec)
        idx_by_g = {g: np.where(test_groups_vec == g)[0] for g in uniq_g}

        # print(clusters)
        # for each cluster: perturb and measure DELTA-RMSE
        for c in clusters:
            # cols = np.where(col_clusters == c)[0]
            cols = np.where(col_clusters == c)[0].astype(np.intp)
            if cols.size == 0:
                continue

            Xp = X_test.copy()

            if conditional:
                # residualized (conditional) perturbation of the whole cluster
                # not_cols = np.setdiff1d(np.arange(d), cols, assume_unique=True)
                not_cols = np.setdiff1d(np.arange(d, dtype=np.intp), cols, assume_unique=True).astype(np.intp)
                if not_cols.size == 0:
                    # print('Conditional')
                    # cannot condition on empty set -> fallback to unconditional permutation
                    for g, loc in idx_by_g.items():
                        if loc.size > 1:
                            for col in cols:
                                perm = rs.permutation(loc.size)
                                Xp[loc, col] = Xp[loc, col][perm]
                else:
                    # print(cols)
                    Ztr = X[inner_train_idx][:, not_cols]
                    Ytr = X[inner_train_idx][:, cols]              # multioutput target
                    Zte = X_test[:, not_cols]
                    zsc = StandardScaler().fit(Ztr)        # train-fold only
                    Ztr_s = zsc.transform(Ztr)
                    Zte_s = zsc.transform(Zte)

                    glm = Ridge(alpha=glm_alpha, fit_intercept=True, random_state=42)
                    glm.fit(Ztr_s, Ytr)                    # targets not scaled
                    MU_te = glm.predict(Zte_s)             # on original units of Ytr
                    R_te  = X_test[:, cols] - MU_te

                    # MU_te = glm.predict(Zte)                        # (n_test, |cols|)
                    # R_te  = X_test[:, cols] - MU_te                 # residuals on TEST

                    # permute residual *rows* within each held-out condition
                    for g, loc in idx_by_g.items():
                        if loc.size <= 1:
                            # no-op: nothing to permute
                            continue
                        perm = rs.permutation(loc.size)
                        R_perm = R_te[loc][perm, :]                 # preserve residual covariance
                        # Xp[loc[:, None], cols] = MU_te[loc, :] + R_perm
                        Xp[np.ix_(loc, cols)] = MU_te[loc, :] + R_perm
            else:
                # classic within-condition permutation per column
                for g, loc in idx_by_g.items():
                    if loc.size > 1:
                        for col in cols:
                            perm = rs.permutation(loc.size)
                            Xp[loc, col] = Xp[loc, col][perm]

            # evaluate perturbed test
            y_pred_p = np.asarray(fold_pipe.predict(Xp)).ravel()
            rmse_p = sqrt(mean_squared_error(y_test, y_pred_p))
            # d = rmse_p - base_rmse
            # dRMSE_sum[c]   += d
            # dRMSE_sumsq[c] += d * d
            delta = rmse_p - base_rmse
            dRMSE_sum[c]   += delta
            dRMSE_sumsq[c] += delta * delta
            counts[c]      += 1
            
            

    # aggregate results
    rows = []
    for c in clusters:
        n = max(1, counts[c])
        mu = dRMSE_sum[c] / n
        var = max(0.0, dRMSE_sumsq[c] / n - mu**2)
        sd  = np.sqrt(var)
        rows.append({"cluster": int(c), "delta_RMSE_mean": mu, "delta_RMSE_sd": sd, "n_iters": counts[c],
                     "conditional": bool(conditional), "glm_alpha": float(glm_alpha)})
    out = pd.DataFrame(rows).sort_values("delta_RMSE_mean", ascending=False).reset_index(drop=True)
    return out, base_list


# ======================================================
# ======================================================
# ======================================================
# ======================================================

## Get Cluster Members
# ================================================
# Per-feature conditional permutation (Ridge) within clusters
# + Excel export (one sheet per cluster, ranked)
# ================================================
# -------------------------
# Core: per-feature DELTA-RMSE within a cluster (Ridge residualization)
# -------------------------
def feature_permutation_importance_within_cluster(
    pipe_template,
    X: np.ndarray, y: np.ndarray,
    cond_keys: np.ndarray,              # grouping vector (length N)
    always_train_keys: Sequence,        # groups that must never appear in TEST
    cluster_assign,                     # Series/array length d (feature -> cluster id)
    target_cluster: int,                # which cluster to rank within
    feature_names: Optional[Sequence[str]] = None,  # names for columns; default "f{idx}"
    n_iters: int = 300, test_k: int = 3, k_valid_groups: int = 3,
    K_AUG: int = 0, rng: int = 42, verbose: bool = True,
    exhaustive: bool = True, shuffle_combos: bool = True,
    conditional_mode: str = "outside_cluster",      # "outside_cluster" or "all_others"
    glm_alpha: float = 1.0,                         # or set to None to CV an alpha grid per replicate
    alpha_grid: Sequence[float] = (0.1, 1.0, 10.0), # used only if glm_alpha is None
    n_repeats: int = 3,                             # repeats for residual-row shuffles within replicate (small)
    skip_singleton_groups: bool = False              # ignore groups with 1 row in TEST when shuffling
):
    """
    Conditional permutation importance for *features in target_cluster* only.
    - Residualize x_j on chosen conditioning set (outside_cluster by default).
    - Shuffle residual *rows* within each held-out condition; recompose x_j := E + shuffled(R).
    - Average DELTA-RMSE across n_repeats within replicate; aggregate across replicates.

    Returns:
      out_df (ranked), base_rmse_per_replicate (array)
    """
    # safety
    X = np.asarray(X); y = np.asarray(y)
    N, d = X.shape
    # if y.ndim > 1 and y.shape[1] > 1:
    #     y = y[:, 0]  # predict first target for RMSE (adapt if multioutput metric desired)
    # y = y.ravel()

    if feature_names is None:
        feature_names = [f"f{i}" for i in range(d)]

    if hasattr(cluster_assign, "values"):
        cl = np.asarray(cluster_assign.values)
    else:
        cl = np.asarray(cluster_assign)
    if cl.shape[0] != d:
        raise ValueError("cluster_assign must have one entry per column in X.")

    feat_idx_in_cluster = np.where(cl == target_cluster)[0]
    if feat_idx_in_cluster.size == 0:
        raise ValueError(f"No features found in cluster {target_cluster}.")

    rs = np.random.RandomState(rng)

    # build test-group selections
    all_groups = np.unique(cond_keys)
    anchor_set = set(always_train_keys)
    non_anchor_groups = np.array([g for g in all_groups if g not in anchor_set])

    from itertools import combinations
    if exhaustive:
        combo_list = [tuple(c) for c in combinations(non_anchor_groups, test_k)]
        if shuffle_combos:
            rs.shuffle(combo_list)
        iter_groups = combo_list
    else:
        iter_groups = [tuple(rs.choice(non_anchor_groups, size=test_k, replace=False)) for _ in range(n_iters)]
    n_iters_eff = len(iter_groups)

    # storage
    idx2name = {i: feature_names[i] for i in range(d)}
    dRMSE_sum   = {j: 0.0 for j in feat_idx_in_cluster}
    dRMSE_sumsq = {j: 0.0 for j in feat_idx_in_cluster}
    counts      = {j: 0   for j in feat_idx_in_cluster}
    base_list   = np.zeros(n_iters_eff, dtype=float)

    # progress bar
    try:
        from tqdm import tqdm
        iterator = tqdm(iter_groups, total=n_iters_eff, disable=not verbose,
                        desc=f"Delta DELTA-RMSE (cluster {target_cluster})")
    except Exception:
        iterator = iter_groups

    # loop replicates
    for it, test_groups in enumerate(iterator):
        test_mask = np.isin(cond_keys, test_groups)
        test_idx  = np.where(test_mask)[0]
        train_idx = np.where(~test_mask)[0]

        # guard: anchors never in test
        if np.isin(np.where(np.isin(cond_keys, list(anchor_set)))[0], test_idx).any():
            raise RuntimeError("Anchor condition leaked into test.")

        # inner valid split (grouped)
        inner_train_idx, inner_valid_idx, _ = make_random_group_valid(
            train_idx, cond_keys, always_train_keys, k=k_valid_groups, rng=42
        )

        # fit model on train
        from skorch.helper import predefined_split
        from skorch.dataset import Dataset
        from skorch.callbacks import EarlyStopping, LRScheduler, GradientNormClipping

        fold_pipe = clone(pipe_template)
        X_train, y_train = X[inner_train_idx], y[inner_train_idx]
        if (K_AUG is not None) and (K_AUG > 1):
            X_train, y_train = augment_labels_normal(
                X_train, y_train, K=K_AUG, rng=42, min_gr=0.0, truncated=True
            )

        if inner_valid_idx is not None and len(inner_valid_idx) > 0:
            X_valid, y_valid = X[inner_valid_idx], y[inner_valid_idx]
            valid_ds = Dataset(X_valid.astype(np.float32, copy=False),
                            y_valid.astype(np.float32, copy=False))
            try:
                fold_pipe.set_params(
                    net__train_split=predefined_split(valid_ds),
                    net__callbacks=[
                        LRScheduler(policy='ReduceLROnPlateau', monitor='valid_loss',
                                    factor=0.5, patience=5, min_lr=1e-6, cooldown=0),
                        GradientNormClipping(1.0),
                        EarlyStopping(patience=50, monitor='valid_loss', load_best=True),
                    ],
                    net__criterion__min_var=1e-4,
                    net__verbose=0,
                    net__iterator_train__shuffle=False,
                    net__iterator_train__num_workers=0,
                    net__iterator_valid__shuffle=False,
                    net__iterator_valid__num_workers=0,
                )
            except Exception:
                pass
        else:
            try:
                fold_pipe.set_params(net__train_split=None, net__callbacks=[], net__verbose=0,
                    net__iterator_train__shuffle=False,
                    net__iterator_train__num_workers=0,
                    net__iterator_valid__shuffle=False,
                    net__iterator_valid__num_workers=0,)
            except Exception:
                pass
        set_seed(42)
        fold_pipe.fit(X_train, y_train)

        # baseline on TEST
        X_test = X[test_idx]
        y_test = y[test_idx]
        if y_test.ndim > 1: y_test = y_test[:, 0]
        y_test = y_test.ravel()
        y_pred = np.asarray(fold_pipe.predict(X_test)).ravel()
        base_rmse = sqrt(mean_squared_error(y_test, y_pred))
        base_list[it] = base_rmse

        # group structure inside TEST
        test_groups_vec = cond_keys[test_idx]
        uniq_g = np.unique(test_groups_vec)
        idx_by_g = {g: np.where(test_groups_vec == g)[0] for g in uniq_g}

        # prepare conditioning sets
        if conditional_mode == "outside_cluster":
            outside_cols = np.where(cl != target_cluster)[0]
            use_outside = outside_cols.size > 0
            if use_outside:
                Ztr = X[inner_train_idx][:, outside_cols]
                Zte = X_test[:, outside_cols]
                zsc_out = StandardScaler().fit(Ztr)
                Ztr_s = zsc_out.transform(Ztr)
                Zte_s = zsc_out.transform(Zte)

        # feature loop (only within cluster)
        for j in feat_idx_in_cluster:
            # choose conditioning matrix per mode
            if conditional_mode == "outside_cluster":
                if not use_outside:
                    # fallback: no conditioning possible -> unconditional within-group permutes of x_j
                    def _resid_rows():
                        return X_test[:, j], np.zeros_like(X_test[:, j])
                else:
                    # optionally CV alpha per replicate/feature (cheap)
                    if glm_alpha is None:
                        best_alpha, best_loss = None, np.inf
                        # simple one-step CV on inner valid
                        Ztr_v = X[inner_train_idx][:, outside_cols]
                        Zvl   = X[inner_valid_idx][:, outside_cols] if inner_valid_idx.size else None
                        zcv = StandardScaler().fit(Ztr_v)
                        Ztr_v_s = zcv.transform(Ztr_v)
                        Zvl_s   = zcv.transform(Zvl) if Zvl is not None else None
                        ytr = X[inner_train_idx][:, j]
                        yvl = X[inner_valid_idx][:, j] if inner_valid_idx.size else None
                        for a in alpha_grid:
                            glm = Ridge(alpha=a, fit_intercept=True, random_state=42)
                            glm.fit(Ztr_v_s, ytr)
                            if Zvl_s is None:
                                # fall back to train reconstruction loss
                                loss = np.mean((ytr - glm.predict(Ztr_v_s))**2)
                            else:
                                loss = np.mean((yvl - glm.predict(Zvl_s))**2)
                            if loss < best_loss:
                                best_loss, best_alpha = loss, a
                        alpha_use = best_alpha if best_alpha is not None else 1.0
                    else:
                        alpha_use = float(glm_alpha)

                    glm = Ridge(alpha=alpha_use, fit_intercept=True, random_state=42)
                    glm.fit(Ztr_s, X[inner_train_idx][:, j])
                    MU_te = glm.predict(Zte_s)                  # E[x_j|outside]
                    R_te  = X_test[:, j] - MU_te

                    def _resid_rows():
                        return MU_te, R_te

            elif conditional_mode == "all_others":
                not_cols = np.setdiff1d(np.arange(d, dtype=int), np.array([j], dtype=int), assume_unique=True)
                if not_cols.size == 0:
                    def _resid_rows():
                        return X_test[:, j], np.zeros_like(X_test[:, j])
                else:
                    Ztr = X[inner_train_idx][:, not_cols]
                    Zte = X_test[:, not_cols]
                    zsc = StandardScaler().fit(Ztr)
                    Ztr_s = zsc.transform(Ztr)
                    Zte_s = zsc.transform(Zte)

                    alpha_use = float(glm_alpha) if glm_alpha is not None else 1.0
                    glm = Ridge(alpha=alpha_use, fit_intercept=True, random_state=42)
                    glm.fit(Ztr_s, X[inner_train_idx][:, j])
                    MU_te = glm.predict(Zte_s)
                    R_te  = X_test[:, j] - MU_te

                    def _resid_rows():
                        return MU_te, R_te
            else:
                raise ValueError("conditional_mode must be 'outside_cluster' or 'all_others'")

            # repeated residual-row shuffles (small n_repeats to smooth tiny groups)
            d_repeats = np.zeros(n_repeats, dtype=float)
            MU_te, R_te = _resid_rows()
            for r in range(n_repeats):
                Xp = X_test.copy()
                for g, loc in idx_by_g.items():
                    if skip_singleton_groups and loc.size <= 1:
                        # leave as-is; contributes no permutation noise
                        Xp[loc, j] = MU_te[loc] + R_te[loc]
                        continue
                    if loc.size > 1:
                        perm = rs.permutation(loc.size)
                        Xp[loc, j] = MU_te[loc] + R_te[loc][perm]
                y_pred_p = np.asarray(fold_pipe.predict(Xp)).ravel()
                rmse_p = sqrt(mean_squared_error(y_test, y_pred_p))
                d_repeats[r] = rmse_p - base_rmse

            d_mean_rep = float(np.mean(d_repeats))
            dRMSE_sum[j]   += d_mean_rep
            dRMSE_sumsq[j] += d_mean_rep * d_mean_rep
            counts[j]      += 1

    # aggregate across replicates
    rows = []
    for j in feat_idx_in_cluster:
        n = max(1, counts[j])
        mu = dRMSE_sum[j] / n
        var = max(0.0, dRMSE_sumsq[j] / n - mu**2)
        sd  = float(np.sqrt(var))
        # CI over replicate means
        ci_lo = ci_hi = np.nan
        if counts[j] > 1 and sd >= 0:
            try:
                from scipy.stats import t
                se = sd / np.sqrt(counts[j])
                tcrit = t.ppf(0.975, df=counts[j]-1)
                ci_lo, ci_hi = mu - tcrit*se, mu + tcrit*se
            except Exception:
                se = sd / np.sqrt(counts[j]); z = 1.96
                ci_lo, ci_hi = mu - z*se, mu + z*se

        rows.append({
            "feature": idx2name[j],
            "feature_index": int(j),
            "cluster": int(target_cluster),
            "delta_RMSE_mean": mu,
            "delta_RMSE_sd": sd,
            "ci95_lo": ci_lo,
            "ci95_hi": ci_hi,
            "n_iters": int(counts[j]),
            "n_repeats": int(n_repeats),
            "conditional_mode": conditional_mode,
            "glm_alpha": float(glm_alpha) if glm_alpha is not None else np.nan
        })

    out = (pd.DataFrame(rows)
           .sort_values("delta_RMSE_mean", ascending=False)
           .assign(rank_in_cluster=lambda d_: np.arange(1, len(d_)+1))
           .reset_index(drop=True))
    # add %DELTA-RMSE relative to median baseline for interpretability
    base_med = float(np.median(base_list)) if len(base_list) else np.nan
    if base_med and base_med > 0:
        out["pct_delta_RMSE"] = 100.0 * out["delta_RMSE_mean"] / base_med
    else:
        out["pct_delta_RMSE"] = np.nan

    return out, base_list

# -------------------------
# Excel writer: one sheet per cluster
# -------------------------
def _pick_excel_engine() -> Optional[str]:
    for eng in ("xlsxwriter", "openpyxl"):
        try:
            __import__(eng)
            return eng
        except Exception:
            continue
    return None

from joblib import Parallel, delayed

def export_cluster_rankings_to_xlsx(
    X: np.ndarray, y: np.ndarray,
    cond_keys: np.ndarray, always_train_keys: Sequence,
    cluster_assign,
    feature_names: Optional[Sequence[str]] = None,
    out_path: str = "per_feature_rankings_by_cluster.xlsx",
    n_iters: int = 300, test_k: int = 3, k_valid_groups: int = 3,
    K_AUG: int = 100, rng: int = 42, exhaustive: bool = True, shuffle_combos: bool = True,
    conditional_mode: str = "outside_cluster",
    glm_alpha: float = 1.0, alpha_grid: Sequence[float] = (0.1, 1.0, 10.0),
    n_repeats: int = 3, skip_singleton_groups: bool = True,
    feature_to_protein_df: Optional[pd.DataFrame] = None,
    acc2kegg_df: Optional[pd.DataFrame] = None,
    pipe_factory: Optional[Callable[[], Any]] = None,   # NEW
    n_jobs: int = 1, backend: str = "loky",
    use_cuda: bool = True, device_ids=None, train_batch_size=None
):
    cl = cluster_assign.values if hasattr(cluster_assign, "values") else np.asarray(cluster_assign)
    d = X.shape[1]
    if cl.shape[0] != d:
        raise ValueError("cluster_assign must have one entry per column in X.")
    if feature_names is None:
        feature_names = [f"f{i}" for i in range(d)]

    protein_map: Dict[Tuple[str,int], str] = {}
    if feature_to_protein_df is not None:
        f2p = feature_to_protein_df.copy()
        need = {"feature","cluster","protein_id"}
        miss = need - set(f2p.columns)
        if miss:
            raise ValueError(f"feature_to_protein_df missing columns: {miss}")
        if acc2kegg_df is not None:
            to_uniprot = _build_to_uniprot_map(acc2kegg_df)
            f2p["protein_id"] = f2p["protein_id"].astype(str).str.strip().map(to_uniprot).fillna(f2p["protein_id"])
        for _, r in f2p.iterrows():
            protein_map[(str(r["feature"]), int(r["cluster"]))] = str(r["protein_id"])

    clusters = sorted(np.unique(cl).astype(int))

    # --- worker used for PROCESS backend: must NOT capture `pipe_template` ---
    def _compute_one_cluster_proc(k: int):
        local_pipe = pipe_factory()  # factory must be top-level & picklable
        # pin device per worker (optional; shown earlier in our GPU patch)
        if use_cuda and device_ids:
            dev_id = device_ids[(k % len(device_ids))]
            try:
                local_pipe.set_params(net__device=f"cuda:{dev_id}")
                if train_batch_size is not None:
                    local_pipe.set_params(net__iterator_train__batch_size=int(train_batch_size))
            except Exception:
                pass
        out_k, base_k = feature_permutation_importance_within_cluster(
            pipe_template=local_pipe, X=X, y=y,
            cond_keys=cond_keys, always_train_keys=always_train_keys,
            cluster_assign=cl, target_cluster=int(k),
            feature_names=feature_names,
            n_iters=n_iters, test_k=test_k, k_valid_groups=k_valid_groups,
            K_AUG=K_AUG, rng=rng, verbose=False,
            exhaustive=exhaustive, shuffle_combos=shuffle_combos,
            conditional_mode=conditional_mode,
            glm_alpha=glm_alpha, alpha_grid=alpha_grid,
            n_repeats=n_repeats, skip_singleton_groups=skip_singleton_groups
        )
        if protein_map:
            out_k["protein_id"] = [protein_map.get((feat, int(k)), "") for feat in out_k["feature"]]
            cols = ["rank_in_cluster","feature","protein_id","cluster","delta_RMSE_mean","delta_RMSE_sd",
                    "ci95_lo","ci95_hi","pct_delta_RMSE","n_iters","n_repeats",
                    "conditional_mode","glm_alpha","feature_index"]
            cols = [c for c in cols if c in out_k.columns] + [c for c in out_k.columns if c not in cols]
            out_k = out_k[cols]
        return (k, out_k, base_k)

    results, errors = [], []
    if n_jobs == 1:
        for k in clusters:
            try: results.append(_compute_one_cluster_proc(k))
            except Exception as e: errors.append({"cluster": int(k), "error": repr(e)})
    else:
        if backend == "loky":
            if pipe_factory is None:
                raise RuntimeError("With process backend ('loky'), pass a top-level picklable pipe_factory().")

            def _safe_proc(k):
                try:
                    return _compute_one_cluster_proc(k)
                except Exception as e:
                    return ("__ERROR__", {"cluster": int(k), "error": repr(e)})

            # --- progress bar (tqdm_joblib if available; else joblib verbose) ---
            try:
                from tqdm.auto import tqdm
                from tqdm_joblib import tqdm_joblib

                with tqdm_joblib(tqdm(total=len(clusters), desc="Clusters", leave=True, dynamic_ncols=True)):
                    par_out = Parallel(n_jobs=n_jobs, backend="loky", prefer="processes")(
                        delayed(_safe_proc)(k) for k in clusters
                    )
            except Exception:
                # Fallback: use joblib's own progress messages
                par_out = Parallel(n_jobs=n_jobs, backend="loky", prefer="processes", verbose=10)(
                    delayed(_safe_proc)(k) for k in clusters
                )
        for item in par_out:
            if item[0] == "__ERROR__": errors.append(item[1])
            else: results.append(item)

    # ... write Excel ..
    engine = _pick_excel_engine()
    writer_kwargs = {"engine": engine} if engine else {}
    with pd.ExcelWriter(out_path, **writer_kwargs) as writer:
        summary = []
        for k, out_k, base_k in results:
            sheet_name = f"cluster_{k}"
            out_k.to_excel(writer, index=False, sheet_name=sheet_name)
            try:
                ws = writer.sheets[sheet_name]
                ws.freeze_panes(1, 0); ws.autofilter(0, 0, len(out_k), len(out_k.columns)-1)
            except Exception:
                pass
            summary.append({
                "cluster": int(k),
                "n_features": len(out_k),
                "baseline_RMSE_median": float(np.median(base_k)) if len(base_k) else np.nan,
                "top_feature": out_k.iloc[0]["feature"] if len(out_k) else "",
                "top_delta_RMSE_mean": out_k.iloc[0]["delta_RMSE_mean"] if len(out_k) else np.nan,
                "top_pct_delta_RMSE": out_k.iloc[0]["pct_delta_RMSE"] if "pct_delta_RMSE" in out_k.columns and len(out_k) else np.nan
            })
        if summary:
            sum_df = pd.DataFrame(summary).sort_values("cluster")
            sum_df.to_excel(writer, index=False, sheet_name="summary")
            try:
                ws = writer.sheets["summary"]
                ws.freeze_panes(1, 0); ws.autofilter(0, 0, len(sum_df), len(sum_df.columns)-1)
            except Exception:
                pass
        if errors:
            pd.DataFrame(errors).to_excel(writer, index=False, sheet_name="errors")

    return os.path.abspath(out_path)



# ==================================
FDR_ALPHA = 0.05
EFFECT_SIGMA_MULT = 2.0
COVERAGE_TARGET = 0.80
USE_PARAMETRIC_BOOT = True
N_BOOT = 200
STABILITY_THRESH = 0.60
RNG_SEED = 42
# ====================================================================

def load_sheet(xlsx_path, sheet_name, required_cols=None, allow_missing=False):
    if not Path(xlsx_path).exists():
        if allow_missing:
            return None
        raise FileNotFoundError(f"File not found: {xlsx_path}")
    try:
        df = pd.read_excel(xlsx_path, sheet_name=sheet_name)
    except Exception as e:
        if allow_missing:
            return None
        raise ValueError(f"Failed reading '{sheet_name}' from '{xlsx_path}': {e}")
    if required_cols is not None:
        lc = {c.lower() for c in df.columns}
        need = {c.lower() for c in required_cols}
        if not need.issubset(lc):
            raise ValueError(
                f"Sheet '{sheet_name}' missing required cols.\nHave: {list(df.columns)}\nNeed: {required_cols}"
            )
        name_map = {}
        for want in required_cols:
            for have in df.columns:
                if have.lower() == want.lower():
                    name_map[have] = want
                    break
        df = df.rename(columns=name_map)
    return df

def attach_uniprot_from_kegg(assign_df, acc2k_df):
    out = assign_df.copy()
    out["feature"] = out["feature"].astype(str).str.strip()
    if acc2k_df is None or acc2k_df.empty:
        out["UniProt"] = np.nan
        out["Description"] = np.nan
        return out
    acc = acc2k_df.copy()
    acc["KEGG"] = acc["KEGG"].astype(str).str.strip()
    acc["UniProt"] = acc["UniProt"].astype(str).str.strip()
    acc = acc.drop_duplicates(subset=["KEGG"], keep="first")
    out = out.merge(acc[["KEGG", "UniProt", "Description"]],
                    left_on="feature", right_on="KEGG", how="left")
    return out.drop(columns=["KEGG"])

def choose_cluster_exemplar_fallback(group):
    if "local_exemplar" in group and group["local_exemplar"].notna().any():
        modes = group["local_exemplar"].dropna().astype(str)
        cnt = Counter(modes)
        top = sorted(cnt.items(), key=lambda x: (-x[1], list(modes).index(x[0])))[0][0]
        return top
    if "subcluster" in group.columns and (group["subcluster"] == 0).any():
        g0 = group[group["subcluster"] == 0]
        if g0["local_exemplar"].notna().any():
            return g0["local_exemplar"].dropna().astype(str).iloc[0]
    return str(group["feature"].iloc[0])

# ---------- Selection helpers (your logic, slightly wrapped) ----------

def bh_fdr(p):
    p = np.asarray(p, float)
    n = p.size
    order = np.argsort(p)
    ranks = np.empty_like(order)
    ranks[order] = np.arange(1, n + 1)
    q = p * n / ranks
    q_sorted = np.minimum.accumulate(q[order][::-1])[::-1]
    q_adj = np.empty_like(q)
    q_adj[order] = np.minimum(q_sorted, 1.0)
    return q_adj

def smallest_k_for_coverage(values_desc, target, total=None):
    v = np.clip(np.asarray(values_desc, float), 0.0, None)
    T = v.sum() if total is None else float(total)
    if T <= 0: return 0
    cum = np.cumsum(v)
    k = int((cum < target * T).sum() + 1)
    return min(k, v.size)

def select_from_summary(df,
                        coverage_target=COVERAGE_TARGET,
                        fdr_alpha=FDR_ALPHA,
                        effect_sigma_mult=EFFECT_SIGMA_MULT,
                        do_boot=USE_PARAMETRIC_BOOT,
                        n_boot=N_BOOT,
                        stab_thr=STABILITY_THRESH,
                        rng_seed=RNG_SEED):
    req = {"cluster","delta_RMSE_mean","delta_RMSE_sd","n_iters"}
    if not req.issubset(df.columns):
        raise ValueError(f"Missing columns. Have {list(df.columns)}, need {sorted(req)}")

    T = df.copy()
    T = T.dropna(subset=["cluster","delta_RMSE_mean","delta_RMSE_sd","n_iters"]).copy()
    T["cluster"] = pd.to_numeric(T["cluster"], errors="coerce").astype("Int64")
    T["n_iters"] = pd.to_numeric(T["n_iters"], errors="coerce")
    T["delta_RMSE_mean"] = pd.to_numeric(T["delta_RMSE_mean"], errors="coerce")
    T["delta_RMSE_sd"]   = pd.to_numeric(T["delta_RMSE_sd"], errors="coerce")
    T = T.sort_values("cluster").reset_index(drop=True)

    mu  = T["delta_RMSE_mean"].values.astype(float)
    sd  = T["delta_RMSE_sd"].values.astype(float)
    n   = np.maximum(1, T["n_iters"].values.astype(int))
    se  = sd / np.sqrt(n)

    try:
        from scipy.stats import norm
    except Exception as e:
        raise ImportError("SciPy is required for normal CDF (scipy.stats.norm). Please install scipy.") from e

    z   = np.divide(mu, se, out=np.zeros_like(mu), where=se>0)
    p   = 1.0 - norm.cdf(z)
    q   = bh_fdr(p)

    effect_thr_mean = effect_sigma_mult * se
    pass_effect = mu > effect_thr_mean
    pass_fdr    = q < fdr_alpha
    pass_both   = pass_fdr & pass_effect

    total_pos = np.clip(mu, 0.0, None).sum()
    idx_pass  = np.where(pass_both)[0]
    if idx_pass.size > 0:
        order_desc = idx_pass[np.argsort(mu[idx_pass])[::-1]]
        mu_pass_desc = mu[order_desc]
        k_cov = smallest_k_for_coverage(mu_pass_desc, coverage_target, total=total_pos)
        keep_by_cov_idx = set(order_desc[:k_cov])
    else:
        k_cov = 0
        keep_by_cov_idx = set()

    stab_freq = None
    final_mask = np.array([i in keep_by_cov_idx for i in range(T.shape[0])]) & pass_both

    if do_boot and idx_pass.size > 0 and n_boot > 0:
        rng = np.random.default_rng(rng_seed)
        sel_counts = np.zeros(T.shape[0], dtype=int)
        for _ in range(n_boot):
            mu_b = rng.normal(loc=mu, scale=se)
            z_b  = np.divide(mu_b, se, out=np.zeros_like(mu_b), where=se>0)
            p_b  = 1.0 - norm.cdf(z_b)
            q_b  = bh_fdr(p_b)
            pass_both_b = (q_b < fdr_alpha) & (mu_b > effect_sigma_mult * se)

            idx_b = np.where(pass_both_b)[0]
            if idx_b.size > 0:
                ord_b = idx_b[np.argsort(mu_b[idx_b])[::-1]]
                mu_b_desc = np.clip(mu_b[ord_b], 0.0, None)
                total_pos_b = np.clip(mu_b, 0.0, None).sum()
                k_b = smallest_k_for_coverage(mu_b_desc, coverage_target, total=total_pos_b)
                final_idx_b = set(ord_b[:k_b])
                for j in final_idx_b:
                    sel_counts[j] += 1

        stab = sel_counts / float(n_boot)
        stab_freq = pd.Series(stab, index=T["cluster"].astype(int).values, name="selection_freq")
        stable_mask = (stab >= stab_thr)
        final_mask = final_mask & stable_mask

    results = pd.DataFrame({
        "cluster": T["cluster"].astype(int),
        "delta_RMSE_mean": mu,
        "delta_RMSE_sd": sd,
        "n_iters": n,
        "SE_mean": se,
        "z_one_sided": z,
        "p_one_sided": p,
        "q_bh": q,
        "effect_thr_mean": effect_thr_mean,
        "pass_effect": pass_effect,
        "pass_fdr": pass_fdr,
        "pass_both": pass_both,
        "kept_by_coverage": np.array([i in keep_by_cov_idx for i in range(T.shape[0])]),
        "stable_ge_thresh": (stab_freq.reindex(T["cluster"].astype(int).values).values >= STABILITY_THRESH) if stab_freq is not None else np.nan,
        "final_selected": final_mask
    }).sort_values("delta_RMSE_mean", ascending=False).reset_index(drop=True)

    selected = results.loc[results["pass_both"], "cluster"].tolist()
    return results, selected, k_cov, stab_freq

# ----------------------- MAIN BUILDER -----------------------
def build_outputs_single_sheet(
    xlsx_path,                      # workbook with: Assignments, cluster_dRMSE_cond, feat_exemplars, shap_by_group
    assign_sheet="Assignments",
    cluster_drmse_sheet="cluster_dRMSE_cond",
    feat_exemplars_sheet="feat_exemplars",
    shap_by_group_sheet="shap_by_group",
    out_xlsx="MergedClusters.xlsx",
    acc2k_xlsx_path="Kegg2Accession.xlsx",
    acc2k_sheet="MasterDict",
):
    """
    Writes FOUR sheets to 'MergedClusters.xlsx' while preserving:
      " the global sort defined on all_df
      " the exact column arrangements requested
    """
    # 1) Load Assignments
    required_assign = ["feature", "parent_cluster", "subcluster", "local_exemplar"]
    df_assign = load_sheet(xlsx_path, assign_sheet, required_assign, allow_missing=False)
    df_assign["parent_cluster"] = pd.to_numeric(df_assign["parent_cluster"], errors="coerce").astype("Int64")
    df_assign["subcluster"]     = pd.to_numeric(df_assign["subcluster"],     errors="coerce").astype("Int64")

    # KEGGÂUniProt/Description mapping
    df_acc2k = load_sheet(
        acc2k_xlsx_path,
        acc2k_sheet,
        required_cols=["KEGG", "UniProt", "Description"],
        allow_missing=False,
    )
    df_members = attach_uniprot_from_kegg(df_assign, df_acc2k)

    # feat_exemplars (use row index as parent_cluster key)
    df_ex = load_sheet(xlsx_path, feat_exemplars_sheet, required_cols=["exemplar"], allow_missing=False).copy()
    try:
        df_ex.index = pd.to_numeric(df_ex.index, errors="coerce").astype("Int64")
    except Exception:
        pass

    def pick_parent_exemplar_by_index(cluster_id, g):
        if pd.notna(cluster_id) and cluster_id in df_ex.index:
            val = df_ex.loc[cluster_id, "exemplar"]
            if pd.notna(val):
                return str(val).strip()
        return choose_cluster_exemplar_fallback(g)

    # Compute parent_exemplar per parent_cluster
    rows = []
    for clus, g in df_members.groupby("parent_cluster", dropna=True, sort=True):
        g = g.copy()
        parent_exemplar = pick_parent_exemplar_by_index(int(clus), g)
        g["parent_exemplar"] = parent_exemplar
        rows.append(g)
    all_df = pd.concat(rows, ignore_index=True) if rows else df_members.assign(parent_exemplar=np.nan)

    # Merge delta_RMSE_mean for global sorting (but DO NOT lose column order later)
    required_drmse = ["cluster", "delta_RMSE_mean", "delta_RMSE_sd", "n_iters", "conditional", "glm_alpha"]
    df_drmse = load_sheet(xlsx_path, cluster_drmse_sheet, required_cols=required_drmse, allow_missing=False).copy()
    df_drmse["cluster"] = pd.to_numeric(df_drmse["cluster"], errors="coerce").astype("Int64")
    df_drmse["delta_RMSE_mean"] = pd.to_numeric(df_drmse["delta_RMSE_mean"], errors="coerce")

    all_df = all_df.merge(
        df_drmse[["cluster", "delta_RMSE_mean"]].rename(columns={"cluster": "parent_cluster",
                                                                 "delta_RMSE_mean": "__delta_RMSE"}),
        on="parent_cluster",
        how="left"
    )

    # Establish the ONE TRUE global order on all_df and keep it for all sheets
    is_sub_ex = (all_df["feature"].astype(str) == all_df["local_exemplar"].astype(str)).astype(int)
    all_df = (
        all_df.assign(__is_sub_exemplar=is_sub_ex)
              .sort_values(
                  by=["__delta_RMSE", "parent_cluster", "subcluster", "__is_sub_exemplar"],
                  ascending=[False, True, True, False]
              )
    )
    # Create a stable feature-order index to reuse later
    feature_order = {feat: i for i, feat in enumerate(all_df["feature"].astype(str).tolist())}

    # Build all_clusters with EXACT column arrangement (no re-sorting beyond all_df order)
    all_df = all_df.rename(columns={"UniProt": "UniProt ID"})
    all_clusters_cols = [
        "feature",
        "UniProt ID",
        "parent_cluster",
        "subcluster",
        "local_exemplar",
        "parent_exemplar",
        "Description",
    ]
    for c in all_clusters_cols:
        if c not in all_df.columns:
            all_df[c] = np.nan
    # Keep the all_df order, then select columns
    all_clusters_df = all_df[all_clusters_cols].copy()

    # shap_by_group (augmented; Description excluded), preserve original cols then appended in fixed order,
    #    AND row order matches all_df via feature_order
    df_shap_in = load_sheet(xlsx_path, shap_by_group_sheet, required_cols=None, allow_missing=False)
    if "feature" not in df_shap_in.columns:
        raise ValueError(f"Input sheet '{shap_by_group_sheet}' must contain a 'feature' column to join on.")
    orig_cols = list(df_shap_in.columns)

    annot_cols = ["UniProt ID", "parent_cluster", "subcluster", "local_exemplar", "parent_exemplar"]
    join_source = all_df[["feature"] + annot_cols].drop_duplicates("feature")
    df_shap_out = df_shap_in.merge(join_source, on="feature", how="left")

    # Preserve original column order, then append new columns in this exact order
    final_shap_cols = orig_cols + [c for c in annot_cols if c not in orig_cols]
    df_shap_out = df_shap_out[final_shap_cols]

    # Row order: align to all_df's order
    df_shap_out["_ord_"] = df_shap_out["feature"].astype(str).map(feature_order)
    df_shap_out = df_shap_out.sort_values("_ord_").drop(columns="_ord_")

    # cluster_selection results using your pipeline
    selection_results, selected_clusters, k_cov, stab_freq = select_from_summary(df_drmse)

    #  shap_by_group_selected filtered to pass_both clusters,
    #     same column order as shap_by_group and SAME row order as all_df
    pass_both_clusters = set(selection_results.loc[selection_results["pass_both"], "cluster"].astype(int).tolist())
    df_shap_selected = df_shap_out[df_shap_out["parent_cluster"].astype("Int64").isin(pass_both_clusters)].copy()

    # Write FOUR sheets (preserving arrangements)
    out_xlsx = Path(out_xlsx)
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as w:
        all_clusters_df.to_excel(w, index=False, sheet_name="all_clusters")
        df_shap_out.to_excel(w, index=False, sheet_name="shap_by_group")
        selection_results.to_excel(w, index=False, sheet_name="cluster_selection")
        df_shap_selected.to_excel(w, index=False, sheet_name="shap_by_group_selected")

    print(
        f"[DONE] wrote '{out_xlsx}' with sheets: "
        f"'all_clusters' ({len(all_clusters_df):,}), "
        f"'shap_by_group' ({len(df_shap_out):,}), "
        f"'cluster_selection' ({len(selection_results):,}), "
        f"'shap_by_group_selected' ({len(df_shap_selected):,}).\n"
        f"Selected (pass_both) clusters: {sorted(pass_both_clusters)}\n"
        f"k for {int(COVERAGE_TARGET*100)}% coverage among passers: {k_cov}"
    )

    return {
        "all_clusters": all_clusters_df,
        "shap_by_group": df_shap_out,
        "cluster_selection": selection_results,
        "shap_by_group_selected": df_shap_selected
    }



# ----------- significance knobs -----------
ALPHA = 0.05         # BH FDR threshold (within each cluster)
EFFECT_MULT = 2.0    # effect-size gate: mu > 2*SE

# ----------- helpers -----------
def norm_feat(s):
    return s.astype(str).str.strip().str.upper()

def read_per_feature_book(path, tag):
    """
    Read a workbook with 1 sheet per cluster ("cluster_XX").
    tag in {"out","all"} to rename delta_RMSE_mean accordingly.
    Returns: feature, cluster, n_iters, d_{tag}, sd_{tag}
    """
    xls = pd.ExcelFile(path)
    rows = []
    for sh in xls.sheet_names:
        if not str(sh).lower().startswith("cluster_"):
            continue
        k = int(str(sh).split("_", 1)[1])
        df = pd.read_excel(xls, sheet_name=sh)
        dfc = df.copy()
        dfc.columns = [c.strip().lower() for c in dfc.columns]
        need = {"feature","delta_rmse_mean","delta_rmse_sd","n_iters"}
        if not need.issubset(dfc.columns):
            missing = need - set(dfc.columns)
            raise ValueError(f"{path}::{sh} missing {sorted(missing)}")
        keep = dfc[["feature","delta_rmse_mean","delta_rmse_sd","n_iters"]].copy()
        keep["feature"] = norm_feat(keep["feature"])
        keep["cluster"] = int(k)
        keep = keep.rename(columns={
            "delta_rmse_mean": f"d_{tag}",
            "delta_rmse_sd":   f"sd_{tag}",
        })
        rows.append(keep)
    if not rows:
        raise ValueError(f"No cluster_* sheets found in {path}")
    out = pd.concat(rows, ignore_index=True)
    # If the two books have different n_iters, keep the max per featureÃcluster
    out["n_iters"] = pd.to_numeric(out["n_iters"], errors="coerce")
    out["n_iters"] = out.groupby(["feature","cluster"])["n_iters"].transform("max")
    return out

def z_in_group(s):
    mu = s.mean()
    sd = s.std(ddof=0)
    return (s - mu) / (sd if sd and np.isfinite(sd) else 1.0)

def bh_fdr_2(p):
    """Benjamini-Hochberg over a 1D array (NaNs -> 1.0)."""
    p = np.asarray(p, float)
    mask = np.isfinite(p)
    q = np.full_like(p, 1.0)
    if mask.sum() == 0:
        return q
    ps = p[mask]
    order = np.argsort(ps)
    ranks = np.empty_like(order)
    ranks[order] = np.arange(1, len(ps)+1)
    q_raw = ps * len(ps) / ranks
    q_sorted = np.minimum.accumulate(q_raw[order][::-1])[::-1]
    q[mask] = np.minimum(q_sorted, 1.0)
    return q

# --- vectorized version ---
def _onesided_p_from_z(z):
    """
    One-sided p = 1 - Phi(z), vectorized.
    Uses numpy's erf if available; otherwise vectorizes math.erf.
    """
    z = np.asarray(z, dtype=float)
    try:
        # numpy exposes a vectorized erf via numpy.special (aliased in top namespace in many builds)
        from numpy import erf as _nerf
    except Exception:
        from math import erf as _erf_scalar
        _nerf = np.vectorize(_erf_scalar, otypes=[float])
    Phi = 0.5 * (1.0 + _nerf(z / np.sqrt(2.0)))
    p = 1.0 - Phi
    return np.clip(p, 0.0, 1.0)


def quad(row, thr=0.0):
    if row["z_out_rank"] >= thr and row["z_all_rank"] >= thr:
        return "HH"
    if row["z_out_rank"] >= thr and row["z_all_rank"] <  thr:
        return "HL"
    if row["z_out_rank"] <  thr and row["z_all_rank"] >= thr:
        return "LH"
    return "LL"


# ----------- per-feature significance (within cluster) for both modes -----------
def add_per_feature_significance_df(df):
    g = df.copy()

    # n iters (rowwise)
    n = np.maximum(1.0, pd.to_numeric(g["n_iters"], errors="coerce").astype(float))

    # ---------- Outside-cluster stats ----------
    mu_out = pd.to_numeric(g["d_out"], errors="coerce")
    sd_out = pd.to_numeric(g["sd_out"], errors="coerce")
    se_out = (sd_out / np.sqrt(n)).replace(0, np.nan)
    zstat_out = mu_out / se_out
    zstat_out = zstat_out.where(np.isfinite(zstat_out),
                                np.where(mu_out > 0, np.inf, -np.inf))
    p_out = pd.Series(_onesided_p_from_z(zstat_out.values), index=g.index)

    # ---------- All-others stats ----------
    mu_all = pd.to_numeric(g["d_all"], errors="coerce")
    sd_all = pd.to_numeric(g["sd_all"], errors="coerce")
    se_all = (sd_all / np.sqrt(n)).replace(0, np.nan)
    zstat_all = mu_all / se_all
    zstat_all = zstat_all.where(np.isfinite(zstat_all),
                                np.where(mu_all > 0, np.inf, -np.inf))
    p_all = pd.Series(_onesided_p_from_z(zstat_all.values), index=g.index)

    # ---------- BH-FDR per cluster (via transform) ----------
    q_out = (
        g.assign(p_out=p_out)
         .groupby("cluster")["p_out"]
         .transform(lambda s: pd.Series(bh_fdr_2(s.to_numpy()), index=s.index))
    )
    q_all = (
        g.assign(p_all=p_all)
         .groupby("cluster")["p_all"]
         .transform(lambda s: pd.Series(bh_fdr_2(s.to_numpy()), index=s.index))
    )

    # ---------- Pass/fail logic ----------
    pass_sign_out   = mu_out > 0
    pass_effect_out = mu_out > (EFFECT_MULT * (se_out.fillna(np.inf)))
    pass_fdr_out    = q_out < ALPHA
    pass_both_out   = pass_sign_out & pass_effect_out & pass_fdr_out

    pass_sign_all   = mu_all > 0
    pass_effect_all = mu_all > (EFFECT_MULT * (se_all.fillna(np.inf)))
    pass_fdr_all    = q_all < ALPHA
    pass_both_all   = pass_sign_all & pass_effect_all & pass_fdr_all

    consensus_pass = pass_both_out & pass_both_all
    consensus_score = pd.Series(
        np.minimum(zstat_out.values, zstat_all.values), index=g.index
    ).replace([np.inf, -np.inf], np.nan)

    # ---------- Assign outputs ----------
    g = g.assign(
        mu_out=mu_out, sd_out=sd_out, se_out=se_out, zstat_out=zstat_out, p_out=p_out, q_out=q_out,
        pass_sign_out=pass_sign_out, pass_effect_out=pass_effect_out,
        pass_fdr_out=pass_fdr_out, pass_both_out=pass_both_out,

        mu_all=mu_all, sd_all=sd_all, se_all=se_all, zstat_all=zstat_all, p_all=p_all, q_all=q_all,
        pass_sign_all=pass_sign_all, pass_effect_all=pass_effect_all,
        pass_fdr_all=pass_fdr_all, pass_both_all=pass_both_all,

        consensus_pass=consensus_pass, consensus_score=consensus_score
    )
    return g



# ----------- ranks WITH SHAP tie-breakers (unchanged) -----------
def compute_ranks_vectorized(paired: pd.DataFrame) -> pd.DataFrame:
    df = paired.copy()

    # ---------- rep_rank ----------
    # sort per-cluster by z_all_rank, z_out_rank, mean_abs_shap (all desc)
    order = df.sort_values(
        ["cluster", "z_all_rank", "z_out_rank", "mean_abs_shap"],
        ascending=[True,     False,        False,          False]
    )
    df.loc[order.index, "rep_rank"] = order.groupby("cluster").cumcount() + 1

    # ---------- proxy_rank ----------
    # proxy_score = z_out_rank - 0.5 * max(z_all_rank, 0)
    df["proxy_score"] = df["z_out_rank"] - 0.5 * np.maximum(df["z_all_rank"], 0.0)
    order = df.sort_values(
        ["cluster", "proxy_score", "mean_abs_shap"],
        ascending=[True,       False,           False]
    )
    df.loc[order.index, "proxy_rank"] = order.groupby("cluster").cumcount() + 1

    # ---------- balanced_rank ----------
    # balanced_score = unique_index - 0.25 * max(redundancy_index, 0)
    df["balanced_score"] = df["unique_index"] - 0.25 * np.maximum(df["redundancy_index"], 0.0)
    order = df.sort_values(
        ["cluster", "balanced_score", "mean_abs_shap"],
        ascending=[True,           False,           False]
    )
    df.loc[order.index, "balanced_rank"] = order.groupby("cluster").cumcount() + 1

    # ---------- final presentational order ----------
    df = df.sort_values(
        ["consensus_pass", "balanced_rank", "rep_rank", "proxy_rank"],
        ascending=[False,            True,         True,         True]
    ).reset_index(drop=True)

    # keep dtypes tidy
    for col in ("rep_rank", "proxy_rank", "balanced_rank"):
        df[col] = df[col].astype(int)

    return df


def rank_cluster(g):
    g = g.copy()

    # Representative (unique) rank: prioritize z_all_rank, then z_out_rank, then mean |SHAP|
    g = g.sort_values(
        by=["z_all_rank","z_out_rank","mean_abs_shap"],
        ascending=[False, False, False]
    )
    g["rep_rank"] = np.arange(1, len(g)+1)

    # Proxy/marker rank: emphasize outside signal but penalize non-unique; SHAP tiebreak
    g["proxy_score"] = g["z_out_rank"] - 0.5*np.maximum(g["z_all_rank"], 0.0)
    g = g.sort_values(
        by=["proxy_score","mean_abs_shap"],
        ascending=[False, False]
    )
    g["proxy_rank"] = np.arange(1, len(g)+1)

    # Balanced rank: both-high first, then SHAP, weaken redundancy
    g["balanced_score"] = g["unique_index"] - 0.25*np.maximum(g["redundancy_index"], 0.0)
    g = g.sort_values(
        by=["balanced_score","mean_abs_shap"],
        ascending=[False, False]
    )
    g["balanced_rank"] = np.arange(1, len(g)+1)

    # Final presentational order: put statistically significant (consensus) on top,
    # then keep your original ranks to break order within blocks.
    g = g.sort_values(["consensus_pass","balanced_rank","rep_rank","proxy_rank"],
                      ascending=[False, True, True, True])
    return g



eps = 1e-12



# Robust testing options
STD_EPS                 = 1e-6    # "near-constant" std threshold
VAR_EPS                 = 1e-12   # tiny variance floor when needed
USE_PERM_IF_CONST       = True    # preferred fallback for const/near-const groups
USE_MWU_IF_CONST        = True    # secondary fallback (median shift)
N_PERM                  = 20000   # permutations (reduce if too slow)
PERM_SEED               = 0

def pick_anchor(df: pd.DataFrame, cond_key: str) -> str:
    if cond_key.endswith("_ae"):
        return "Ac_ae"
    if cond_key.endswith("_an"):
        return "Ac_an"
    sub = df[df["Condition_key"] == cond_key]
    if len(sub) == 0:
        raise ValueError(f"Unknown condition key: {cond_key}")
    oxy_vals = sub["Oxygen"].dropna().unique().tolist()
    if len(oxy_vals) != 1:
        raise ValueError(f"Ambiguous Oxygen for {cond_key}: {oxy_vals}")
    return "Ac_ae" if str(oxy_vals[0]).lower().startswith("aerob") else "Ac_an"


# ---------------- ----------------
def compute_log2fc(group_vals: np.ndarray, ctrl_vals: np.ndarray, log2_already: bool) -> float:
    g = group_vals[~np.isnan(group_vals)]
    c = ctrl_vals[~np.isnan(ctrl_vals)]
    if g.size == 0 or c.size == 0:
        return np.nan
    if log2_already:
        return float(np.nanmean(g) - np.nanmean(c))
    else:
        eps = 1e-12
        return float(np.log2((np.nanmean(g) + eps) / (np.nanmean(c) + eps)))


def perm_test_mean_diff(a: np.ndarray, b: np.ndarray, n_perm=10000, seed=0) -> float:
    a = a[~np.isnan(a)]
    b = b[~np.isnan(b)]
    if a.size < 2 or b.size < 2:
        return np.nan
    rng = np.random.default_rng(seed)
    obs = float(np.mean(a) - np.mean(b))
    pool = np.concatenate([a, b])
    n_a = a.size
    count = 0
    for _ in range(n_perm):
        perm = rng.permutation(pool)
        diff = float(np.mean(perm[:n_a]) - np.mean(perm[n_a:]))
        if abs(diff) >= abs(obs):
            count += 1
    return (count + 1) / (n_perm + 1)


def robust_pval(g_vals: np.ndarray, c_vals: np.ndarray) -> tuple[float, str]:
    g = g_vals[~np.isnan(g_vals)]
    c = c_vals[~np.isnan(c_vals)]
    if g.size < 2 or c.size < 2:
        return np.nan, "insufficient_n"

    g_mean, c_mean = float(np.mean(g)), float(np.mean(c))
    g_std, c_std = float(np.std(g, ddof=1)), float(np.std(c, ddof=1))

    both_const = (g_std < STD_EPS) and (c_std < STD_EPS)
    one_const  = (g_std < STD_EPS) ^ (c_std < STD_EPS)

    if both_const:
        return 1.0, "both_const_p1"

    if one_const:
        if USE_PERM_IF_CONST:
            try:
                p = perm_test_mean_diff(g, c, n_perm=N_PERM, seed=PERM_SEED)
                return float(p), "perm_mean_diff"
            except Exception:
                pass
        if USE_MWU_IF_CONST:
            try:
                _, p = mannwhitneyu(g, c, alternative="two-sided", method="auto")
                return float(p), "mannwhitney_u"
            except Exception:
                pass
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            _, p = ttest_ind_from_stats(g_mean, max(g_std, VAR_EPS), g.size,
                                        c_mean, max(c_std, VAR_EPS), c.size,
                                        equal_var=False)
        return float(p), "welch_var_floor"

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        _, p = ttest_ind(g, c, equal_var=False)
    if not np.isfinite(p):
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            _, p = ttest_ind_from_stats(g_mean, max(g_std, VAR_EPS), g.size,
                                        c_mean, max(c_std, VAR_EPS), c.size,
                                        equal_var=False)
        return float(p), "welch_from_stats"
    return float(p), "welch"


def volcano_plot(
    df_sub: pd.DataFrame,
    cond_label: str,
    *,
    alpha=0.05,
    l2fc_thr=1.0,
    outdir=Path("."),
    label_map: dict | None = None,
    annotate_all_sig: bool = False,
    top_k_sig_labels: int = 50,
    min_abs_log2fc_for_label: float = 0.0
) -> Path:
    """Generic volcano; annotate significant points."""
    x = df_sub["log2FC"].to_numpy()
    p = df_sub["pval"].to_numpy()
    yplot = -np.log10(np.clip(p, 1e-300, 1.0))

    sig_mask = (df_sub["adj_pval"] < alpha) & (np.abs(df_sub["log2FC"]) >= l2fc_thr)

    fig = plt.figure(figsize=(7.0, 5.5), dpi=140)
    ax = plt.gca()
    ax.scatter(x[~sig_mask], yplot[~sig_mask], s=6, alpha=0.5)
    ax.scatter(x[sig_mask], yplot[sig_mask], s=12, alpha=0.9)

    ax.axhline(-np.log10(alpha), linestyle="--")
    ax.axvline(l2fc_thr, linestyle="--")
    ax.axvline(-l2fc_thr, linestyle="--")

    ax.set_xlabel("log2 Fold Change")
    ax.set_ylabel("-log10 p-value")
    ax.set_title(f"Volcano: {cond_label}")
    ax.xaxis.set_major_locator(MaxNLocator(nbins=6))
    ax.yaxis.set_major_locator(MaxNLocator(nbins=6))

    # annotate significant
    ann_df = df_sub[sig_mask].copy()
    if min_abs_log2fc_for_label > 0:
        ann_df = ann_df[ann_df["log2FC"].abs() >= min_abs_log2fc_for_label]
    if not ann_df.empty:
        if not annotate_all_sig:
            ann_df = ann_df.nsmallest(top_k_sig_labels, "pval")
        for _, row in ann_df.iterrows():
            pid = row["Protein_ID"]
            lbl = label_map.get(pid, pid) if label_map else pid
            ax.annotate(
                lbl,
                xy=(row["log2FC"], -np.log10(max(row["pval"], 1e-300))),
                xytext=(2, 2),
                textcoords="offset points",
                fontsize=7
            )

    fig.tight_layout()
    safe_label = "".join(ch if ch.isalnum() or ch in "._-+" else "_" for ch in cond_label)
    outpath = outdir / f"volcano__{safe_label}.png"
    fig.savefig(outpath)
    plt.close(fig)
    return outpath


def heatmap_fc(matrix: pd.DataFrame, title: str, outname: str, vmax=3.0, vmin=-3.0, outdir=Path(".")) -> Path:
    if matrix.shape[0] == 0 or matrix.shape[1] == 0:
        fig = plt.figure(figsize=(6, 3), dpi=140)
        ax = plt.gca()
        ax.text(0.5, 0.5, "No data to display", ha="center", va="center")
        ax.axis("off")
        outpath = outdir / f"{outname}.png"
        fig.savefig(outpath)
        plt.close(fig)
        return outpath

    M = matrix.copy().to_numpy(dtype=float)
    M = np.clip(M, vmin, vmax)

    # fig = plt.figure(figsize=(max(6, 0.25*matrix.shape[1]+2), max(5, 0.18*matrix.shape[0]+2)), dpi=140)
    # height scaling stronger (0.25 * nrows + 2)
    fig_h = max(5, 0.25 * matrix.shape[0] + 2)
    fig_w = max(6, 0.25 * matrix.shape[1] + 2)
    fig = plt.figure(figsize=(fig_w, fig_h), dpi=140)
    ax = plt.gca()
    im = ax.imshow(M, aspect="auto", interpolation="nearest")
    ax.set_title(title)
    ax.set_xlabel("Comparison" if outname.lower().startswith("o2") else "Condition")
    ax.set_ylabel("Protein_ID")
    ax.set_xticks(np.arange(matrix.shape[1]))
    ax.set_xticklabels(matrix.columns, rotation=45, ha="right")
    ax.set_yticks(np.arange(matrix.shape[0]))
    ax.set_yticklabels(matrix.index)
    cbar = plt.colorbar(im)
    cbar.set_label("log2FC")
    fig.tight_layout()

    outpath = outdir / f"{outname}.png"
    fig.savefig(outpath, bbox_inches="tight")
    plt.close(fig)
    return outpath

def format_excel_with_filters_and_freeze(path):
    wb = load_workbook(path)
    for ws in wb.worksheets:
        # Determine the full data range
        max_row = ws.max_row
        max_col = ws.max_column
        if max_col == 0 or max_row == 0:
            continue
        # A1:Z1000 style range for autofilter
        data_range = f"A1:{ws.cell(row=max_row, column=max_col).coordinate}"
        # Apply filter dropdowns
        ws.auto_filter.ref = data_range
        # Freeze the top row: freezes everything above row 2 (so row 1 stays visible)
        ws.freeze_panes = "A2"
    wb.save(path)
    print(f"Formatted workbook: {path}")
